import logging
import sqlite3
import time
import json
import os
import asyncio
from functools import partial
from datetime import datetime, timedelta, time as dt_time

# Try to import openpyxl, guide user if not installed
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("Error: 'openpyxl' library not found. Please install it using: pip install openpyxl")
    exit()

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove, InputFile
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ConversationHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
from telegram.constants import ParseMode
from telegram.error import BadRequest

# --- بخش تنظیمات و ثابت‌ها ---
BOT_TOKEN = "7598913970:AAG3zD7svL1xKRpb-VDnYgCf4cBS7aKqlTk" # <<<< توکن ربات خود را اینجا قرار دهید
OWNER_ID = 7487830899 # <<<< شناسه کاربری ادمین (خودتان) را اینجا قرار دهید
BANK_CARD_NUMBER = "۶۰۳۷-۹۹۷۹-۹۹۹۹-۹۹۹۹"
SUPPORT_PHONE = "09123456789"
SUPPORT_USERNAME = "@YourSupportUsername" # <<<< آیدی پشتیبانی را اینجا وارد کنید
TRADER_BOT_USERNAME = "@MyAwesomeTraderBot"
AI_BOT_USERNAME = "@soawmygptbot" # <<<< نام کاربری ربات هوش مصنوعی اصلی
CRM_BOT_PRODUCT_NAME = "بات crm"
EXCEL_FILE_NAME = "orders.xlsx"

# --- توضیحات محصولات (این بخش را به راحتی ویرایش کنید) ---
PRODUCT_DESCRIPTIONS = {
    "بات تریدر": "این یک ربات معامله‌گر پیشرفته برای بازارهای مالی است که با استفاده از الگوریتم‌های هوشمند، به صورت خودکار برای شما معامله می‌کند.",
    "بات هوش مصنوعی": "یک دستیار هوش مصنوعی قدرتمند که می‌تواند در تولید محتوا، پاسخگویی به سوالات پیچیده و اتوماسیون وظایف روزمره به شما کمک کند. برای استفاده از قابلیت‌های پیشرفته‌تر، می‌توانید اعتبار تهیه کنید.",
    "بات crm": "یک ربات جامع برای مدیریت ارتباط با مشتری (CRM) که به شما امکان ثبت اطلاعات مشتریان، پیگیری سفارش‌ها و ارسال پیام‌های خودکار را می‌دهد.",
    "چت بات": "ما برای پلتفرم‌های مختلف (تلگرام، دیسکورد، وب‌سایت و...) چت‌بات‌های هوشمند و سفارشی طراحی می‌کنیم. لطفاً پلتفرم مورد نظر خود را انتخاب کنید.",
    "اعتبار چت جی پی تی 4": "بسته اعتباری برای فعال‌سازی و استفاده از مدل قدرتمند GPT-4 در ربات هوش مصنوعی شما.",
    "اعتبار چت جی پی تی 5": "بسته اعتباری ویژه برای دسترسی به پیشرفته‌ترین مدل هوش مصنوعی، GPT-5، با قابلیت‌های منحصر به فرد."
}

# --- راه‌اندازی لاگین ---
logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logging.getLogger("httpx").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


# --- order logging helpers ---
def _ensure_logs_dir():
    try:
        os.makedirs(os.path.join("logs", "orders"), exist_ok=True)
    except Exception:
        logger.exception("Failed to create logs directory")


def log_order_event(order_id: str, text: str):
    """Append a timestamped event line to logs/orders/{order_id}.txt"""
    try:
        _ensure_logs_dir()
        path = os.path.join("logs", "orders", f"{order_id}.txt")
        with open(path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().isoformat()}] {text}\n")
    except Exception:
        logger.exception("Failed to write order log for %s", order_id)


def log_general(text: str):
    """Append general bot events to logs/general.log"""
    try:
        os.makedirs("logs", exist_ok=True)
        path = os.path.join("logs", "general.log")
        with open(path, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().isoformat()}] {text}\n")
    except Exception:
        logger.exception("Failed to write general log")


def log_order_snapshot(order_id: str):
    """Write a snapshot of the order row, user profile and active services into the order log file."""
    try:
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM orders WHERE order_id = ?", (order_id,))
        row = cursor.fetchone()
        cursor.execute("SELECT id, product_name, expiry_date, is_active FROM active_services WHERE user_id = (SELECT user_id FROM orders WHERE order_id = ?)", (order_id,))
        services = cursor.fetchall()
        conn.close()

        lines = []
        lines.append("=== ORDER SNAPSHOT ===")
        try:
            lines.append(json.dumps(row, ensure_ascii=False, default=str))
        except Exception:
            lines.append(str(row))
        lines.append("=== ACTIVE SERVICES ===")
        for s in services:
            try:
                lines.append(json.dumps(s, ensure_ascii=False, default=str))
            except Exception:
                lines.append(str(s))
        lines.append("=== END SNAPSHOT ===")

        _ensure_logs_dir()
        path = os.path.join("logs", "orders", f"{order_id}.txt")
        with open(path, "a", encoding="utf-8") as f:
            f.write("\n".join([f"[{datetime.now().isoformat()}] {l}" for l in lines]) + "\n")
    except Exception:
        logger.exception("Failed to write order snapshot for %s", order_id)


# --- تعریف وضعیت‌های مکالمه ---
(
    MAIN_MENU, SELECTING_PRODUCT, SHOWING_DETAILS, WAITING_FOR_RECEIPT,
    ASK_CRM_LOGO, ASK_CRM_SHOP_NAME, ASK_PRODUCT_INFO_METHOD, AWAITING_PRODUCT_FILE, ASK_CRM_PRODUCT_COUNT,
    ASK_CRM_PRODUCT_NAMES, ASK_CRM_PRICES, ASK_CRM_SHIPPING_QUERY, ASK_CRM_CARD_NUMBER, ASK_CRM_FULL_NAME,
    ASK_CRM_PHONE, WAITING_FOR_RENEWAL_RECEIPT, RENEW_CRM_CHOICE, AWAITING_RENEW_CRM_FILE,
    ASK_FOR_WHOM, GET_TARGET_USER_ID,
    AI_CREDIT_MENU, GET_CREDIT_BUYER_NAME, GET_CREDIT_BUYER_PHONE, WAITING_FOR_CREDIT_RECEIPT,
    CHATBOT_PLATFORM, CHATBOT_HAS_BOT, CHATBOT_GET_TOKEN, CHATBOT_GET_DESC, CHATBOT_GET_SITE_INFO,
    WAITING_FOR_CHATBOT_RECEIPT,
    # --- وضعیت‌های جدید برای کاستوم بات هوش مصنوعی ---
    CUSTOM_AI_START, CUSTOM_AI_ASK_FILE, CUSTOM_AI_GET_FILE, CUSTOM_AI_GET_DESC,
    CUSTOM_AI_GET_NAME, CUSTOM_AI_GET_PHONE, WAITING_FOR_CUSTOM_AI_RECEIPT,
    USER_REJECT_REASON_PROMPT, USER_REJECT_GET_REASON
) = range(39)


# --- توابع دیتابیس ---
def setup_database():
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS products (id INTEGER PRIMARY KEY, name TEXT NOT NULL, price INTEGER NOT NULL)")
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS orders (
        order_id TEXT PRIMARY KEY, user_id INTEGER NOT NULL, user_username TEXT,
        products_json TEXT NOT NULL, total_price INTEGER NOT NULL,
        status TEXT NOT NULL DEFAULT 'pending_approval',
        receipt_file_id TEXT, full_name TEXT, phone TEXT,
        crm_details_json TEXT, order_type TEXT DEFAULT 'new_purchase', 
        related_service_id INTEGER, target_user_id INTEGER,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS active_services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        product_id INTEGER NOT NULL,
        product_name TEXT NOT NULL,
        expiry_date DATE NOT NULL,
        is_active BOOLEAN DEFAULT 1,
        FOREIGN KEY (product_id) REFERENCES products (id)
    )""")
    
    try: cursor.execute("ALTER TABLE orders ADD COLUMN crm_details_json TEXT")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE orders ADD COLUMN order_type TEXT DEFAULT 'new_purchase'")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE orders ADD COLUMN related_service_id INTEGER")
    except sqlite3.OperationalError: pass
    try: cursor.execute("ALTER TABLE orders ADD COLUMN target_user_id INTEGER")
    except sqlite3.OperationalError: pass
        
    cursor.execute("DELETE FROM products")
    sample_products = [
        ("بات تریدر", 200000), 
        ("بات هوش مصنوعی", 0),
        (CRM_BOT_PRODUCT_NAME, 500000),
        ("اعتبار چت جی پی تی 4", 150000),
        ("اعتبار چت جی پی تی 5", 300000),
        ("چت بات", 0),
        ("کاستوم بات هوش مصنوعی", 0) # محصول جدید برای ثبت در دیتابیس
    ]
    cursor.executemany("INSERT INTO products (name, price) VALUES (?, ?)", sample_products)
    conn.commit()
    conn.close()
    logger.info("پایگاه داده با محصولات جدید راه‌اندازی شد.")

# --- توابع اکسل ---
def setup_excel_file():
    if not os.path.exists(EXCEL_FILE_NAME):
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        
        headers = ["نام", "نام خانوادگی", "شماره تلفن", "سرویس‌ها", "شماره سفارش", "وضعیت سفارش", "روز های مانده", "برای کاربر"]
        sheet = workbook.create_sheet(title="All Orders")
        sheet.append(headers)
        
        workbook.save(EXCEL_FILE_NAME)
        logger.info(f"فایل اکسل '{EXCEL_FILE_NAME}' با شیت‌ لازم ساخته شد.")

def log_to_excel(full_name, phone, services_str, order_id, status, remaining_days, target_user_id=None):
    try:
        workbook = load_workbook(EXCEL_FILE_NAME)
        sheet = workbook["All Orders"]
        
        name_parts = full_name.split(" ", 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ""
        
        target_user_str = str(target_user_id) if target_user_id else "خودش"
        
        remaining_days_str = str(remaining_days) if remaining_days is not None else "N/A"

        new_row = [first_name, last_name, phone, services_str, order_id, status, remaining_days_str, target_user_str]
        sheet.append(new_row)
        workbook.save(EXCEL_FILE_NAME)
        logger.info(f"سفارش {order_id} در فایل اکسل ثبت شد.")
    except Exception as e:
        logger.error(f"خطا در ثبت اطلاعات در اکسل برای سفارش {order_id}: {e}")

def update_excel_status(order_id, new_status):
    try:
        workbook = load_workbook(EXCEL_FILE_NAME)
        sheet = workbook["All Orders"]
        for row in sheet.iter_rows(min_row=2):
            if row[4].value == order_id:
                row[5].value = new_status
                if new_status == "رد شده":
                    row[6].value = 0
                workbook.save(EXCEL_FILE_NAME)
                logger.info(f"وضعیت سفارش {order_id} در اکسل به '{new_status}' تغییر کرد.")
                return
    except Exception as e:
        logger.error(f"خطا در به‌روزرسانی وضعیت در اکسل برای سفارش {order_id}: {e}")

# --- توابع کمکی ---
def persian_format_number(num): return f"{num:,}".replace(",", "،")
def generate_order_id(): return f"SHOP-{int(time.time())}"

# --- توابع اصلی و منوها ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    context.user_data.clear()
    
    keyboard = [
        [InlineKeyboardButton("🛍️ مشاهده محصولات", callback_data="view_products")],
        [InlineKeyboardButton("👤 سرویس‌های من", callback_data="my_services")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    message_content = f"سلام {user.mention_html()} عزیز! 👋\nبه ربات ما خوش آمدید. لطفاً یک گزینه را انتخاب کنید:"
    
    if update.callback_query:
        await update.callback_query.message.edit_text(message_content, reply_markup=reply_markup, parse_mode=ParseMode.HTML)
    else:
        await update.message.reply_html(message_content, reply_markup=reply_markup)
        
    return MAIN_MENU

async def back_to_products(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    if query: await query.answer()
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    # Exclude custom AI bot from main product list as it has its own entry point
    cursor.execute("SELECT id, name FROM products WHERE name NOT LIKE 'اعتبار چت جی پی تی%' AND name != 'کاستوم بات هوش مصنوعی'")
    products = cursor.fetchall()
    conn.close()
    
    keyboard = []
    for product_id, name in products:
        text = f"🔹 {name}"
        keyboard.append([InlineKeyboardButton(text, callback_data=f"details_{product_id}")])
    
    if 'cart' not in context.user_data:
        context.user_data['cart'] = {}
        
    keyboard.append([InlineKeyboardButton("🛒 مشاهده سبد خرید", callback_data="view_cart")])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="main_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    message_text = "برای خرید اشتراک ماهیانه، یکی از سرویس‌های زیر را انتخاب کنید:"
    await query.edit_message_text(message_text, reply_markup=reply_markup)
        
    return SELECTING_PRODUCT

async def show_product_details(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    product_id = int(query.data.split("_")[1])
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT name, price FROM products WHERE id = ?", (product_id,))
    product = cursor.fetchone()
    conn.close()

    if not product:
        await query.edit_message_text("خطا: محصول یافت نشد.")
        return SELECTING_PRODUCT

    name, price = product
    description = PRODUCT_DESCRIPTIONS.get(name, "توضیحات برای این محصول موجود نیست.")
    
    if name == "بات هوش مصنوعی":
        text = f"**{name}**\n\n{description}"
        keyboard = [
            [InlineKeyboardButton("🤖 ورود به ربات هوش مصنوعی", url=f"https://t.me/{AI_BOT_USERNAME.lstrip('@')}")],
            [InlineKeyboardButton("🎁 بات هوش مصنوعی رایگان", url="https://t.me/soawmyvpnbot")],
            [InlineKeyboardButton("💎 خرید اعتبار", callback_data="buy_credits_ai")],
            [InlineKeyboardButton("✨ کاستوم بات هوش مصنوعی", callback_data="custom_ai_start")], # <<<< دکمه جدید
            [InlineKeyboardButton("🔙 بازگشت به لیست محصولات", callback_data="back_to_products")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
        return SHOWING_DETAILS
    
    if name == "چت بات":
        text = f"**{name}**\n\n{description}"
        keyboard = [
            [InlineKeyboardButton("🤖 چت بات تلگرام", callback_data="chatbot_telegram")],
            [InlineKeyboardButton("👾 چت بات دیسکورد", callback_data="chatbot_discord")],
            [InlineKeyboardButton("🌐 چت بات سایت", callback_data="chatbot_site")],
            [InlineKeyboardButton("📜 سورس کد چت بات", callback_data="chatbot_source")],
            [InlineKeyboardButton("🧩 سایر", callback_data="chatbot_other")],
            [InlineKeyboardButton("🔙 بازگشت به لیست محصولات", callback_data="back_to_products")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
        return CHATBOT_PLATFORM
    
    text = (f"**{name}**\n\n{description}\n\n**💰 قیمت اشتراک یک ماهه:** {persian_format_number(price)} تومان")
    
    keyboard = [
        [InlineKeyboardButton("➕ افزودن به سبد خرید", callback_data=f"add_{product_id}")],
    ]

    if name == "بات تریدر":
        keyboard.append([InlineKeyboardButton("🤖 تست بات", url="t.me/soatradertestbot")])

    keyboard.extend([
        [InlineKeyboardButton("🛒 مشاهده سبد خرید", callback_data="view_cart")],
        [InlineKeyboardButton("🔙 بازگشت به لیست محصولات", callback_data="back_to_products")]
    ])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
    return SHOWING_DETAILS

# --- مکالمه جدید برای کاستوم بات هوش مصنوعی ---
async def custom_ai_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    context.user_data.clear() # Clear previous data
    context.user_data['custom_ai_request'] = {}

    text = (
        "شما بخش **کاستوم بات هوش مصنوعی** را انتخاب کرده‌اید.\n\n"
        "در این بخش می‌توانید یک ربات هوش مصنوعی کاملاً شخصی‌سازی شده بر اساس نیازها یا دانش اختصاصی خودتان (مثلاً یک فایل PDF یا وب‌سایت) سفارش دهید.\n\n"
        "لطفاً نوع درخواست خود را انتخاب کنید:"
    )
    keyboard = [
        [InlineKeyboardButton("🤖 کاستوم بات هوش مصنوعی (رایگان)", callback_data="custom_ai_type_free")],
        [InlineKeyboardButton("✨ کاستوم بات هوش مصنوعی (پیشرفته)", callback_data="custom_ai_type_premium")],
        [InlineKeyboardButton("🔙 بازگشت", callback_data=f"details_{get_product_id_by_name('بات هوش مصنوعی')}")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
    return CUSTOM_AI_START

async def custom_ai_ask_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    request_type = query.data.split('_')[-1]
    context.user_data['custom_ai_request']['type'] = request_type

    text = "آیا فایلی (مانند PDF, Word, Txt) دارید که می‌خواهید ربات بر اساس اطلاعات آن آموزش ببیند؟"
    keyboard = [
        [InlineKeyboardButton("✅ بله، فایل دارم", callback_data="custom_ai_has_file_yes")],
        [InlineKeyboardButton("❌ خیر، فایل ندارم", callback_data="custom_ai_has_file_no")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.edit_message_text(text, reply_markup=reply_markup)
    return CUSTOM_AI_ASK_FILE

async def custom_ai_get_file_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data.split('_')[-1]

    if choice == 'yes':
        await query.edit_message_text("بسیار خب. لطفاً فایل خود را ارسال کنید.")
        return CUSTOM_AI_GET_FILE
    else: # no
        context.user_data['custom_ai_request']['file_id'] = None
        await query.edit_message_text("متوجه شدم.\n\nلطفاً **توضیحات کامل** در مورد قابلیت‌ها و دانش مورد نظر خود برای ربات را در یک پیام ارسال کنید:", parse_mode=ParseMode.MARKDOWN)
        return CUSTOM_AI_GET_DESC

async def custom_ai_get_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not update.message.document:
        await update.message.reply_text("خطا: لطفاً یک فایل (Document) ارسال کنید.")
        return CUSTOM_AI_GET_FILE
    
    context.user_data['custom_ai_request']['file_id'] = update.message.document.file_id
    await update.message.reply_text(
        "فایل شما دریافت شد. ✅\n\n"
        "اکنون لطفاً **توضیحات کامل** در مورد قابلیت‌ها و دانش مورد نظر خود برای ربات را در یک پیام ارسال کنید:",
        parse_mode=ParseMode.MARKDOWN
    )
    return CUSTOM_AI_GET_DESC

async def custom_ai_get_desc(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['custom_ai_request']['description'] = update.message.text
    await update.message.reply_text("توضیحات شما دریافت شد.\n\nبرای ثبت درخواست، لطفاً **نام و نام خانوادگی** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return CUSTOM_AI_GET_NAME

async def custom_ai_get_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['custom_ai_request']['full_name'] = update.message.text
    await update.message.reply_text("متشکرم. در آخر، **شماره تماس** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return CUSTOM_AI_GET_PHONE

async def custom_ai_finalize_request(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    request_data = context.user_data['custom_ai_request']
    request_data['phone'] = update.message.text
    
    order_id = generate_order_id()
    product_name = f"کاستوم بات AI ({request_data['type']})"
    products_json = json.dumps([{"name": product_name}], ensure_ascii=False)
    details_json = json.dumps(request_data, ensure_ascii=False)

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO orders (order_id, user_id, user_username, products_json, total_price, crm_details_json, order_type, status, full_name, phone)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (order_id, user.id, user.username, products_json, 0, details_json, 'custom_ai_request', 'pending_admin_approval', request_data['full_name'], request_data['phone']))
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"درخواست شما با شماره پیگیری `{order_id}` با موفقیت ثبت شد. ✅\n"
        "ادمین به زودی درخواست شما را بررسی کرده و نتیجه را برایتان ارسال خواهد کرد.",
        parse_mode=ParseMode.MARKDOWN
    )

    admin_message = (f"🔔 **درخواست جدید (کاستوم بات AI)** 🔔\n\n"
                     f"شماره سفارش: `{order_id}`\n"
                     f"کاربر: {user.mention_html()} (ID: `{user.id}`)\n"
                     f"نام: {request_data['full_name']}\n"
                     f"تلفن: {request_data['phone']}\n"
                     f"نوع درخواست: **{request_data['type']}**\n\n"
                     f"**توضیحات کاربر:**\n{request_data['description']}")

    if request_data.get('file_id'):
        await context.bot.send_document(OWNER_ID, request_data['file_id'], caption=admin_message, parse_mode=ParseMode.HTML)
    else:
        await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.HTML)

    keyboard_admin = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ تأیید اولیه و قیمت‌گذاری", callback_data=f"admin_approve_custom_{order_id}"),
        InlineKeyboardButton("❌ رد درخواست", callback_data=f"admin_reject_custom_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, "لطفاً درخواست را بررسی و اقدام کنید:", reply_markup=keyboard_admin)
    
    context.user_data.clear()
    return ConversationHandler.END

# ...existing code...
# ...existing code...
async def receive_custom_ai_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """
    دفاعی‌تر و قابل اطمینان‌تر: این نسخه دریافت فیش برای سفارش‌های کاستوم-AI را پایش می‌کند،
    شناسه سفارش را از context یا در صورت نبودن از دیتابیس بازیابی می‌کند، فایل فیش را
    ذخیره و وضعیت سفارش را آپدیت می‌کند و سپس به شکل مطمئن پیام + فیش را به ادمین ارسال می‌کند.
    """
    user = update.effective_user
    order_id = context.user_data.get('order_id_for_receipt')

    # fallback: اگر order_id در context نبود، از دیتابیس آخرین سفارش مناسب کاربر را پیدا کن
    if not order_id:
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT order_id FROM orders 
            WHERE user_id = ? AND status IN ('user_approved','pending_user_approval','pending_approval')
            ORDER BY timestamp DESC LIMIT 1
        """, (user.id,))
        row = cursor.fetchone()
        conn.close()
        if row:
            order_id = row[0]

    if not order_id:
        await update.message.reply_text("خطای داخلی رخ داد. سفارش پیدا نشد. لطفاً با پشتیبانی تماس بگیرید.")
        return ConversationHandler.END

    # قبول هم عکس و هم فایل (اگر کاربر فیش را به عنوان document فرستاد)
    receipt_file_id = None
    is_photo = False
    if update.message.photo:
        receipt_file_id = update.message.photo[-1].file_id
        is_photo = True
    elif update.message.document:
        receipt_file_id = update.message.document.file_id
    else:
        await update.message.reply_text("لطفا فقط عکس یا فایل فیش واریزی را بفرستید.")
        return WAITING_FOR_CUSTOM_AI_RECEIPT

    # به‌روز رسانی سفارش در دیتابیس
    try:
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("UPDATE orders SET status = ?, receipt_file_id = ? WHERE order_id = ?", ('pending_approval', receipt_file_id, order_id))
        conn.commit()

        cursor.execute("SELECT user_id, user_username, total_price, full_name, phone, crm_details_json FROM orders WHERE order_id = ?", (order_id,))
        result = cursor.fetchone()
        conn.close()
    except Exception as e:
        logger.error(f"DB error while updating receipt for order {order_id}: {e}")
        await update.message.reply_text("خطا در ثبت فیش در سرور. لطفاً دوباره تلاش کنید یا با پشتیبانی تماس بگیرید.")
        return ConversationHandler.END

    if not result:
        await update.message.reply_text("خطا: سفارش یافت نشد.")
        return ConversationHandler.END

    user_id_db, username_db, price, full_name, phone, crm_details_str = result

    await update.message.reply_text(
        f"فیش شما برای سفارش `{order_id}` دریافت شد. ✅\nپس از تایید نهایی توسط ادمین، کار آغاز خواهد شد.",
        parse_mode=ParseMode.MARKDOWN
    )

    # آماده‌سازی پیام ادمین (HTML برای پایداری بیشتر)
    try:
        details = json.loads(crm_details_str) if crm_details_str else {}
    except Exception:
        details = {}

    platform = details.get('type', 'کاستوم AI') if isinstance(details, dict) else 'کاستوم AI'
    desc = details.get('description') or details.get('details') or 'توضیحات ثبت نشده'

    admin_message = (
        f"🔔 <b>تایید پرداخت (کاستوم بات AI)</b> 🔔\n\n"
        f"📌 <b>شماره سفارش:</b> <code>{order_id}</code>\n"
        f"👤 <b>کاربر:</b> @{username_db if username_db else 'N/A'} (ID: <code>{user_id_db}</code>)\n"
        f"📛 <b>نام:</b> {full_name or 'N/A'}\n"
        f"📞 <b>تلفن:</b> {phone or 'N/A'}\n"
        f"💰 <b>مبلغ:</b> {persian_format_number(price)} تومان\n\n"
        f"📝 <b>توضیحات کاربر:</b>\n{desc}\n\n"
        f"👇 فیش و فایل‌های سفارش:"
    )

    # ارسال امن پیام و فایل به ادمین
    try:
        await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.HTML)
    except Exception as e:
        logger.error(f"Failed to send admin text for order {order_id}: {e}")
        # تلاش مجدد با متن ساده
        try:
            await context.bot.send_message(OWNER_ID, admin_message)
        except Exception as e2:
            logger.error(f"Second attempt to notify admin failed for order {order_id}: {e2}")

    # ارسال هر فایل پیوست کاربر (مثلاً فایل آموزش یا لوگو) اگر موجود باشد
    if isinstance(details, dict):
        # فایل محصول/آموزش
        file_id = details.get('file_id') or details.get('product_info_file_id')
        if file_id:
            try:
                await context.bot.send_document(OWNER_ID, file_id, caption=f"فایل سفارش `{order_id}`")
            except Exception as e:
                logger.error(f"Failed to send attached document for order {order_id} to admin: {e}")

        # لوگو
        logo_id = details.get('logo_file_id')
        if logo_id and logo_id != 'skipped':
            try:
                await context.bot.send_photo(OWNER_ID, logo_id, caption="لوگوی مشتری")
            except Exception as e:
                logger.error(f"Failed to send logo for order {order_id} to admin: {e}")

    # ارسال فیش (عکس یا فایل) با fallback
    try:
        if is_photo:
            await context.bot.send_photo(OWNER_ID, receipt_file_id, caption=f"فیش واریزی برای سفارش `{order_id}`")
        else:
            await context.bot.send_document(OWNER_ID, receipt_file_id, caption=f"فیش واریزی برای سفارش `{order_id}`")
    except Exception as e:
        logger.error(f"Failed to send receipt for order {order_id} to admin: {e}")
        # در صورتی که ارسال فایل ناموفق بود، حداقل شناسه فایل را ارسال کن
        try:
            await context.bot.send_message(OWNER_ID, f"شناسه فیش برای سفارش <code>{order_id}</code>: <code>{receipt_file_id}</code>", parse_mode=ParseMode.HTML)
        except Exception as e2:
            logger.error(f"Also failed to send fallback receipt id for {order_id}: {e2}")

    # دکمه‌های ادمین (تأیید/رد)
    keyboard_admin = InlineKeyboardMarkup([[ 
        InlineKeyboardButton("✅ تایید نهایی پرداخت", callback_data=f"admin_confirm_{order_id}"),
        InlineKeyboardButton("❌ رد پرداخت", callback_data=f"admin_reject_{order_id}")
    ]])
    try:
        await context.bot.send_message(OWNER_ID, "لطفا پرداخت را تایید یا رد کنید:", reply_markup=keyboard_admin)
    except Exception as e:
        logger.error(f"Failed to send admin action keyboard for order {order_id}: {e}")

    # پاکسازی context امن
    context.user_data.pop('order_id_for_receipt', None)
    # مطمئن شو کاربر هم کلیدهای موقت را ندارد
    for k in ('custom_ai_request', 'receipt_file_id'):
        context.user_data.pop(k, None)

    return ConversationHandler.END
# ...existing code...

async def handle_chatbot_payment_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    order_id = context.user_data.get('order_id_for_receipt')

    # fallback: اگر context پاک شده، آخرین سفارش کاربر با وضعیت مناسب را از DB بگیر
    if not order_id:
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT order_id FROM orders
            WHERE user_id = ? AND status IN ('user_approved','pending_user_approval','pending_approval')
            ORDER BY timestamp DESC LIMIT 1
        """, (user.id,))
        row = cursor.fetchone()
        conn.close()
        if row:
            order_id = row[0]

    if not order_id:
        logger.warning("handle_chatbot_payment_receipt: order_id not found for user %s", user.id)
        await update.message.reply_text("خطایی رخ داد. لطفاً دوباره از طریق دکمه تایید اقدام کنید یا با پشتیبانی تماس بگیرید.")
        return ConversationHandler.END

    if not update.message.photo:
        await update.message.reply_text("لطفا فقط عکس فیش واریزی را بفرستید.")
        return WAITING_FOR_CHATBOT_RECEIPT

    receipt_file_id = update.message.photo[-1].file_id

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET status = ?, receipt_file_id = ? WHERE order_id = ?", ('pending_approval', receipt_file_id, order_id))
    conn.commit()
    try:
        log_order_event(order_id, f"RECEIPT RECEIVED file_id={receipt_file_id} by @{user.username} (user_id={user.id})")
        log_order_snapshot(order_id)
    except Exception:
        logger.exception("Failed to log receipt for %s", order_id)

    cursor.execute("SELECT user_id, user_username, total_price, crm_details_json, full_name, phone FROM orders WHERE order_id = ?", (order_id,))
    result = cursor.fetchone()
    if not result:
        logger.error(f"Could not find order {order_id} after receiving payment receipt.")
        conn.close()
        return ConversationHandler.END

    user_id, username, price, details_json_str, full_name, phone = result
    conn.close()

    await update.message.reply_text(
        f"فیش شما برای سفارش `{order_id}` دریافت شد. ✅\nپس از تایید نهایی توسط ادمین، کار ساخت ربات شما آغاز خواهد شد.",
        parse_mode=ParseMode.MARKDOWN
    )

    # آماده‌سازی پیام ادمین و ارسال
    try:
        details = json.loads(details_json_str) if details_json_str else {}
    except Exception:
        details = {}

    platform = details.get('platform', 'N/A').title() if isinstance(details, dict) else 'N/A'
    description = details.get('description', 'توضیحات ثبت نشده') if isinstance(details, dict) else 'N/A'

    admin_message = (f"🔔 **تایید پرداخت (سفارش چت بات)** 🔔\n\n"
                     f"شماره سفارش: `{order_id}`\n"
                     f"کاربر: @{username} (ID: `{user_id}`)\n"
                     f"**پلتفرم:** {platform}\n"
                     f"**مبلغ:** {persian_format_number(price)} تومان\n\n"
                     f"**توضیحات اولیه کاربر:**\n{description}\n\n"
                     "👇 فیش واریزی:")

    await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.MARKDOWN)
    await context.bot.send_photo(OWNER_ID, receipt_file_id)

    keyboard_admin = InlineKeyboardMarkup([[ 
        InlineKeyboardButton("✅ تایید نهایی پرداخت", callback_data=f"admin_confirm_{order_id}"),
        InlineKeyboardButton("❌ رد پرداخت", callback_data=f"admin_reject_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, "لطفا پرداخت را تایید یا رد کنید:", reply_markup=keyboard_admin)

    context.user_data.pop('order_id_for_receipt', None)
    return ConversationHandler.END
# ...existing code...

# --- مکالمه چت‌بات ---
async def chatbot_platform_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    platform = query.data.split("_")[1]
    
    context.user_data['chatbot_request'] = {'platform': platform}

    if platform in ['telegram', 'discord']:
        keyboard = [
            [InlineKeyboardButton("بله، دارم", callback_data="has_bot_yes")],
            [InlineKeyboardButton("خیر، برایم بسازید", callback_data="has_bot_no")]
        ]
        await query.edit_message_text(f"شما **چت بات {platform.title()}** را انتخاب کردید.\n\nآیا از قبل ربات خود را در {platform.title()} ساخته‌اید و توکن آن را دارید؟", reply_markup=InlineKeyboardMarkup(keyboard))
        return CHATBOT_HAS_BOT
    elif platform == 'site':
        await query.edit_message_text("شما **چت بات سایت** را انتخاب کردید.\n\nلطفاً **لینک وب‌سایت** خود را ارسال کنید:")
        return CHATBOT_GET_SITE_INFO
    else: # سایر یا سورس کد
        await query.edit_message_text(f"شما **{platform.replace('_', ' ')}** را انتخاب کردید.\n\nلطفاً **توضیحات کامل** درخواست خود را در یک پیام ارسال کنید:")
        return CHATBOT_GET_DESC

async def chatbot_has_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data.split("_")[2]

    if choice == 'yes':
        await query.edit_message_text("بسیار خب. لطفاً **توکن ربات** خود را ارسال کنید:")
        return CHATBOT_GET_TOKEN
    else: # no
        context.user_data['chatbot_request']['token'] = 'needs_creation'
        await query.edit_message_text("متوجه شدم. ما ربات را برای شما خواهیم ساخت.\n\nاکنون لطفاً **توضیحات کامل** در مورد قابلیت‌های مورد نظر خود را در یک پیام ارسال کنید:")
        return CHATBOT_GET_DESC

async def chatbot_get_token(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['chatbot_request']['token'] = update.message.text
    await update.message.reply_text("توکن دریافت شد. ✅\n\nاکنون لطفاً **توضیحات کامل** در مورد قابلیت‌های مورد نظر خود را در یک پیام ارسال کنید:")
    return CHATBOT_GET_DESC

async def chatbot_get_site_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['chatbot_request']['site_link'] = update.message.text
    await update.message.reply_text("لینک سایت دریافت شد. ✅\n\nاکنون لطفاً **توضیحات کامل** در مورد قابلیت‌های مورد نظر و اطلاعاتی که می‌خواهید ربات داشته باشد را در یک پیام ارسال کنید:")
    return CHATBOT_GET_DESC

async def chatbot_get_desc(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['chatbot_request']['description'] = update.message.text
    await update.message.reply_text("توضیحات شما دریافت شد. در حال ثبت درخواست...")
    return await finalize_chatbot_request(update, context)

async def finalize_chatbot_request(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    order_id = generate_order_id()
    request_data = context.user_data['chatbot_request']
    
    details_json = json.dumps(request_data, ensure_ascii=False)
    product_name = f"چت بات {request_data['platform'].title()}"
    products_json = json.dumps([{"name": product_name}], ensure_ascii=False)

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO orders (order_id, user_id, user_username, products_json, total_price, crm_details_json, order_type, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (order_id, user.id, user.username, products_json, 0, details_json, 'chatbot_request', 'pending_quote'))
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"درخواست شما با شماره پیگیری `{order_id}` با موفقیت ثبت شد. ✅\n"
        "ادمین به زودی درخواست شما را بررسی کرده و هزینه و جزئیات را برایتان ارسال خواهد کرد.",
        parse_mode=ParseMode.MARKDOWN
    )

    admin_message = (f"🔔 **درخواست جدید (سفارش چت بات)** 🔔\n\n"
                     f"شماره سفارش: `{order_id}`\n"
                     f"کاربر: {user.mention_html()} (ID: `{user.id}`)\n"
                     f"**پلتفرم:** {request_data.get('platform', 'N/A').title()}\n")
    if 'token' in request_data:
        admin_message += f"**توکن:** `{request_data['token']}`\n"
    if 'site_link' in request_data:
        admin_message += f"**لینک سایت:** {request_data['site_link']}\n"
    admin_message += f"\n**توضیحات کاربر:**\n{request_data.get('description', 'N/A')}"

    keyboard_admin = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ ارائه قیمت و توضیحات", callback_data=f"admin_quote_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.HTML, reply_markup=keyboard_admin)
    
    context.user_data.clear()
    return ConversationHandler.END

# --- توابع مدیریت قیمت‌گذاری و تایید کاربر ---
async def handle_chatbot_quote_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    order_id = "_".join(query.data.split("_")[2:])

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT total_price FROM orders WHERE order_id = ?", (order_id,))
    result = cursor.fetchone()
    if not result:
        await query.edit_message_text("خطا: سفارش یافت نشد.")
        conn.close()
        return ConversationHandler.END
    
    price = result[0]
    cursor.execute("UPDATE orders SET status = ? WHERE order_id = ?", ('user_approved', order_id))
    conn.commit()
    conn.close()

    context.user_data['order_id_for_receipt'] = order_id

    await query.edit_message_text(
        f"پیشنهاد تایید شد. لطفاً مبلغ **{persian_format_number(price)} تومان** را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
        parse_mode=ParseMode.MARKDOWN
    )
    return WAITING_FOR_CHATBOT_RECEIPT

async def handle_chatbot_quote_reject(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    order_id = "_".join(query.data.split("_")[2:])

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET status = ? WHERE order_id = ?", ('rejected_by_user', order_id))
    conn.commit()
    conn.close()
    
    await query.edit_message_text("پیشنهاد رد شد. درخواست شما لغو گردید.")
    await context.bot.send_message(OWNER_ID, f"❌ کاربر پیشنهاد برای سفارش `{order_id}` را رد کرد.")
    return ConversationHandler.END

async def handle_chatbot_payment_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    order_id = context.user_data.get('order_id_for_receipt')
    if not order_id:
        logger.warning("handle_chatbot_payment_receipt triggered but order_id not in user_data.")
        await update.message.reply_text("خطایی رخ داد. لطفاً دوباره از طریق دکمه تایید اقدام کنید یا با پشتیبانی تماس بگیرید.")
        return ConversationHandler.END

    if not update.message.photo:
        await update.message.reply_text("لطفا فقط عکس فیش واریزی را بفرستید.")
        return WAITING_FOR_CHATBOT_RECEIPT

    receipt_file_id = update.message.photo[-1].file_id
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET status = ?, receipt_file_id = ? WHERE order_id = ?", ('pending_approval', receipt_file_id, order_id))
    conn.commit()
    
    cursor.execute("SELECT user_id, user_username, total_price, crm_details_json FROM orders WHERE order_id = ?", (order_id,))
    result = cursor.fetchone()
    if not result:
        logger.error(f"Could not find order {order_id} after receiving payment receipt.")
        conn.close()
        return ConversationHandler.END

    user_id, username, price, details_json_str = result
    conn.close()

    await update.message.reply_text(
        f"فیش شما برای سفارش `{order_id}` دریافت شد. ✅\nپس از تایید نهایی توسط ادمین، کار ساخت ربات شما آغاز خواهد شد.",
        parse_mode=ParseMode.MARKDOWN
    )
    
    details = json.loads(details_json_str)
    platform = details.get('platform', 'N/A').title()
    description = details.get('description', 'توضیحات ثبت نشده')
    
    admin_message = (f"🔔 **تایید پرداخت (سفارش چت بات)** 🔔\n\n"
                     f"شماره سفارش: `{order_id}`\n"
                     f"کاربر: @{username} (ID: `{user_id}`)\n"
                     f"**پلتفرم:** {platform}\n"
                     f"**مبلغ:** {persian_format_number(price)} تومان\n\n"
                     f"**توضیحات اولیه کاربر:**\n{description}\n\n"
                     "👇 فیش واریزی:")
                     
    await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.MARKDOWN)
    await context.bot.send_photo(OWNER_ID, receipt_file_id)
    
    keyboard_admin = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ تایید نهایی پرداخت", callback_data=f"admin_confirm_{order_id}"),
        InlineKeyboardButton("❌ رد پرداخت", callback_data=f"admin_reject_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, "لطفا پرداخت را تایید یا رد کنید:", reply_markup=keyboard_admin)

    if 'order_id_for_receipt' in context.user_data:
        del context.user_data['order_id_for_receipt']
        
    return ConversationHandler.END

# --- توابع خرید اعتبار هوش مصنوعی ---
async def prompt_ai_credit_options(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, price FROM products WHERE name LIKE 'اعتبار چت جی پی تی%'")
    credit_products = cursor.fetchall()
    conn.close()

    text = "برای استفاده از مدل‌های پیشرفته‌تر، لطفاً بسته اعتباری مورد نظر خود را انتخاب کنید:\n\n"
    keyboard = []
    for prod_id, name, price in credit_products:
        text += f"🔹 **{name}**\n - قیمت: {persian_format_number(price)} تومان\n\n"
        keyboard.append([InlineKeyboardButton(f"خرید {name}", callback_data=f"select_credit_{prod_id}")])

    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data=f"details_{get_product_id_by_name('بات هوش مصنوعی')}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
    return AI_CREDIT_MENU

def get_product_id_by_name(name):
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM products WHERE name = ?", (name,))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None

async def start_credit_purchase(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    product_id = int(query.data.split("_")[2])
    
    context.user_data['credit_product_id'] = product_id
    
    await query.edit_message_text("برای ثبت سفارش اعتبار، لطفاً **نام و نام خانوادگی** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return GET_CREDIT_BUYER_NAME

async def get_credit_buyer_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['full_name'] = update.message.text
    await update.message.reply_text("متشکرم. اکنون **شماره تماس** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return GET_CREDIT_BUYER_PHONE

async def get_credit_buyer_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['phone'] = update.message.text
    product_id = context.user_data['credit_product_id']

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT price FROM products WHERE id = ?", (product_id,))
    price = cursor.fetchone()[0]
    conn.close()

    context.user_data['total_price'] = price
    
    await update.message.reply_text(
        f"اطلاعات شما ثبت شد. لطفاً مبلغ **{persian_format_number(price)} تومان** را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
        parse_mode=ParseMode.MARKDOWN
    )
    return WAITING_FOR_CREDIT_RECEIPT

async def receive_credit_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not update.message.photo:
        await update.message.reply_text("لطفا فقط عکس فیش واریزی را بفرستید.")
        return WAITING_FOR_CREDIT_RECEIPT
    
    receipt_file_id = update.message.photo[-1].file_id
    context.user_data['receipt_file_id'] = receipt_file_id
    
    user_data = context.user_data
    user = update.effective_user
    order_id = generate_order_id()
    product_id = user_data['credit_product_id']

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT name, price FROM products WHERE id = ?", (product_id,))
    name, price = cursor.fetchone()
    
    product_details = [{"id": product_id, "name": name, "price": price, "quantity": 1}]
    products_json = json.dumps(product_details, ensure_ascii=False)

    cursor.execute("""
        INSERT INTO orders (order_id, user_id, user_username, products_json, total_price, receipt_file_id, full_name, phone, order_type)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (order_id, user.id, user.username, products_json, price, receipt_file_id,
          user_data['full_name'], user_data['phone'], 'ai_credit'))
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"سفارش شما برای خرید اعتبار با موفقیت ثبت شد! 🎉\nشماره سفارش شما: `{order_id}`\n"
        "پس از تایید مدیر، اعتبار به ربات شما اضافه خواهد شد.",
        parse_mode=ParseMode.MARKDOWN
    )

    admin_message = (f"🔔 **سفارش جدید (خرید اعتبار AI)** 🔔\n\n"
                     f"شماره سفارش: `{order_id}`\n"
                     f"خریدار: {user.mention_html()} (ID: `{user.id}`)\n"
                     f"نام خریدار: {user_data['full_name']}\n"
                     f"تلفن خریدار: {user_data['phone']}\n\n"
                     f"**محصول:** {name}\n"
                     f"**مبلغ:** {persian_format_number(price)} تومان\n\n"
                     "👇 فیش واریزی:")
    
    await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.HTML)
    await context.bot.send_photo(OWNER_ID, receipt_file_id)
    
    keyboard_admin = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ تایید سفارش", callback_data=f"admin_confirm_{order_id}"),
        InlineKeyboardButton("❌ رد سفارش", callback_data=f"admin_reject_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, "لطفا سفارش را تایید یا رد کنید:", reply_markup=keyboard_admin)
    
    context.user_data.clear()
    return ConversationHandler.END

# --- توابع سبد خرید و پرداخت ---
async def product_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    product_id = int(query.data.split("_")[1])
    
    if 'cart' not in context.user_data:
        context.user_data['cart'] = {}
        
    cart = context.user_data['cart']
    cart[product_id] = cart.get(product_id, 0) + 1
    context.user_data['cart'] = cart
    
    await query.answer("✅ محصول به سبد خرید اضافه شد!", show_alert=True)
    return await back_to_products(update, context)

async def view_cart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    if query: await query.answer()
    
    cart = context.user_data.get('cart', {})
    if not cart:
        text = "سبد خرید شما خالی است."
        keyboard = [[InlineKeyboardButton("🔙 بازگشت به لیست محصولات", callback_data="back_to_products")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        if query: await query.edit_message_text(text, reply_markup=reply_markup)
        else: await update.effective_message.reply_text(text, reply_markup=reply_markup)
        return SELECTING_PRODUCT
        
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    
    cart_items_text = "🛒 **سبد خرید شما:**\n\n"
    total_price = 0
    keyboard = []
    
    for product_id, quantity in cart.items():
        cursor.execute("SELECT name, price FROM products WHERE id = ?", (product_id,))
        product = cursor.fetchone()
        if product:
            name, price = product
            item_total = price * quantity
            total_price += item_total
            cart_items_text += f"▪️ {name} (تعداد: {quantity}) - {persian_format_number(item_total)} تومان\n"
            keyboard.append([InlineKeyboardButton(f"➖ ۱ عدد از «{name}»", callback_data=f"remove_{product_id}")])
    conn.close()

    context.user_data['total_price'] = total_price
    cart_items_text += f"\n\n💳 **مبلغ نهایی قابل پرداخت: {persian_format_number(total_price)} تومان**"
    
    keyboard.extend([
        [InlineKeyboardButton("💳 ادامه و پرداخت", callback_data="checkout")],
        [InlineKeyboardButton("🔙 بازگشت به لیست محصولات", callback_data="back_to_products")]
    ])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(text=cart_items_text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
    else:
        await update.message.reply_text(text=cart_items_text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)

    return SELECTING_PRODUCT

async def remove_from_cart(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    product_id = int(query.data.split("_")[1])
    cart = context.user_data.get('cart', {})
    if product_id in cart:
        cart[product_id] -= 1
        if cart[product_id] <= 0:
            del cart[product_id]
    context.user_data['cart'] = cart
    return await view_cart(update, context)

async def checkout(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    cart = context.user_data.get('cart', {})
    if not cart:
        await query.edit_message_text("سبد خرید شما خالی است!")
        return SELECTING_PRODUCT
        
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    
    crm_in_cart = False
    for product_id in cart.keys():
        cursor.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = cursor.fetchone()[0]
        if product_name == CRM_BOT_PRODUCT_NAME:
            crm_in_cart = True
            break
    conn.close()

    if not crm_in_cart:
        keyboard = [
            [InlineKeyboardButton("بله، برای خودم", callback_data="for_self")],
            [InlineKeyboardButton("خیر، برای شخص دیگری", callback_data="for_other")]
        ]
        await query.edit_message_text("آیا این سرویس(ها) را برای خودتان می‌خواهید؟", reply_markup=InlineKeyboardMarkup(keyboard))
        return ASK_FOR_WHOM
    else:
        await query.edit_message_text(
            "برای تکمیل خرید، لطفاً ابتدا **نام و نام خانوادگی** خود را وارد کنید:",
            parse_mode=ParseMode.MARKDOWN
        )
        return ASK_CRM_LOGO

async def ask_for_whom(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data

    if choice == "for_other":
        await query.edit_message_text("لطفاً **یوزر آیدی عددی** شخص مورد نظر را وارد کنید:")
        return GET_TARGET_USER_ID
    else: # for_self
        context.user_data['target_user_id'] = None
        await query.edit_message_text(
            "بسیار خب. برای ادامه، لطفاً **نام و نام خانوادگی** خود را وارد کنید:",
            parse_mode=ParseMode.MARKDOWN
        )
        return ASK_CRM_LOGO

async def get_target_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    target_id = update.message.text
    if not target_id.isdigit():
        await update.message.reply_text("خطا: لطفاً فقط یوزر آیدی عددی را وارد کنید.")
        return GET_TARGET_USER_ID
    
    context.user_data['target_user_id'] = int(target_id)
    await update.message.reply_text(
        "یوزر آیدی ثبت شد. اکنون لطفاً **نام و نام خانوادگی خودتان** را (به عنوان خریدار) وارد کنید:",
        parse_mode=ParseMode.MARKDOWN
    )
    return ASK_CRM_LOGO

async def receive_name_for_order(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['full_name'] = update.message.text
    await update.message.reply_text("متشکرم. اکنون **شماره تماس** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return ASK_CRM_SHOP_NAME

async def receive_phone_for_order(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['phone'] = update.message.text
    
    cart = context.user_data.get('cart', {})
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    
    crm_in_cart = False
    for product_id in cart.keys():
        cursor.execute("SELECT name FROM products WHERE id = ?", (product_id,))
        product_name = cursor.fetchone()[0]
        if product_name == CRM_BOT_PRODUCT_NAME:
            crm_in_cart = True
            break
    conn.close()
    
    if crm_in_cart:
        context.user_data['crm_details'] = {}
        await update.message.reply_text("عالی. چون سفارش شما شامل بات CRM است، لطفا اطلاعات فروشگاه خود را وارد کنید. ابتدا **لوگوی فروشگاه** را ارسال کنید (یا /skip بزنید):", parse_mode=ParseMode.MARKDOWN)
        return ASK_PRODUCT_INFO_METHOD
    else:
        total_price = context.user_data.get('total_price', 0)
        await update.message.reply_text(
            f"اطلاعات شما ثبت شد. لطفاً مبلغ **{persian_format_number(total_price)} تومان** را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
            parse_mode=ParseMode.MARKDOWN
        )
        return WAITING_FOR_RECEIPT

async def ask_crm_logo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if update.message.photo: context.user_data['crm_details']['logo_file_id'] = update.message.photo[-1].file_id
    elif update.message.text == '/skip': context.user_data['crm_details']['logo_file_id'] = 'skipped'
    else:
        await update.message.reply_text("لطفا یک عکس به عنوان لوگو ارسال کنید یا از دستور /skip استفاده کنید.")
        return ASK_PRODUCT_INFO_METHOD
    await update.message.reply_text("ممنون. حالا لطفاً **اسم فروشگاه** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return AWAITING_PRODUCT_FILE

async def ask_crm_shop_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['crm_details']['shop_name'] = update.message.text
    keyboard = [
        [InlineKeyboardButton("ارسال به صورت فایل (PDF, Word, ...)", callback_data="send_file")],
        [InlineKeyboardButton("ورود دستی اطلاعات در ربات", callback_data="manual_entry")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "عالی! حالا اطلاعات محصولات خود را (تعداد، اسامی، قیمت‌ها) چگونه می‌خواهید ارسال کنید؟",
        reply_markup=reply_markup
    )
    return ASK_CRM_PRODUCT_COUNT

async def ask_product_info_method(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data
    if choice == 'send_file':
        await query.edit_message_text("بسیار خب. لطفاً فایل خود (PDF, Word, Txt و ...) را که شامل اطلاعات محصولات است، ارسال کنید.")
        return ASK_CRM_PRODUCT_NAMES
    else: # manual_entry
        await query.edit_message_text("متوجه شدم. لطفاً به سوالات زیر پاسخ دهید:\n\n**تعداد تقریبی محصولات** فروشگاه شما چندتاست؟ (فقط عدد)", parse_mode=ParseMode.MARKDOWN)
        return ASK_CRM_PRICES

async def awaiting_product_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not update.message.document:
        await update.message.reply_text("لطفاً یک فایل ارسال کنید.")
        return ASK_CRM_PRODUCT_NAMES
    document = update.message.document
    context.user_data['crm_details']['product_info_file_id'] = document.file_id
    context.user_data['crm_details']['product_info_file_name'] = document.file_name
    await update.message.reply_text(
        "فایل شما با موفقیت دریافت شد. ✅\n\n"
        "آیا فروشگاه شما محصولات فیزیکی دارد و نیاز به **هزینه ارسال** دارد؟ (بله / خیر)",
        parse_mode=ParseMode.MARKDOWN
    )
    return ASK_CRM_SHIPPING_QUERY

async def ask_crm_product_count(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['crm_details']['product_count'] = update.message.text
    await update.message.reply_text("حالا لطفاً **اسامی محصولات** را در یک پیام، با ویرگول (,) جدا کرده و بفرستید:", parse_mode=ParseMode.MARKDOWN)
    return ASK_CRM_CARD_NUMBER

async def ask_crm_product_names(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['crm_details']['product_names'] = update.message.text
    await update.message.reply_text("بسیار خب. لطفاً **لیست قیمت‌ها** را متناسب با اسامی محصولات وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return ASK_CRM_FULL_NAME

async def ask_crm_prices(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['crm_details']['prices'] = update.message.text
    await update.message.reply_text("آیا فروشگاه شما محصولات فیزیکی دارد و نیاز به **هزینه ارسال** دارد؟ (بله / خیر)", parse_mode=ParseMode.MARKDOWN)
    return ASK_CRM_SHIPPING_QUERY

async def ask_crm_shipping_query(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['crm_details']['has_shipping'] = 'بله' in update.message.text.lower()
    await update.message.reply_text("متوجه شدم. حالا **شماره کارت بانکی** جهت واریز وجه توسط مشتریان را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
    return ASK_CRM_PHONE

async def ask_crm_card_number(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data['crm_details']['card_number'] = update.message.text
    total_price = context.user_data.get('total_price', 0)
    await update.message.reply_text(
        f"اطلاعات شما ثبت شد. لطفاً مبلغ **{persian_format_number(total_price)} تومان** را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
        parse_mode=ParseMode.MARKDOWN
    )
    return WAITING_FOR_RECEIPT

async def receive_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not update.message.photo:
        await update.message.reply_text("لطفا فقط عکس فیش واریزی رو بفرست.")
        return WAITING_FOR_RECEIPT
    receipt_photo = update.message.photo[-1]
    context.user_data['receipt_file_id'] = receipt_photo.file_id
    await update.message.reply_text("ممنونم! فیش شما دریافت شد. در حال ثبت نهایی سفارش شما...")
    return await finalize_order(update, context)

async def finalize_order(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_data = context.user_data
    user = update.effective_user
    order_id = generate_order_id()
    
    cart = user_data.get('cart', {})
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()

    product_details = []
    for product_id, quantity in cart.items():
        cursor.execute("SELECT name, price FROM products WHERE id = ?", (product_id,))
        name, price = cursor.fetchone()
        product_details.append({"id": product_id, "name": name, "price": price, "quantity": quantity})

    products_json = json.dumps(product_details, ensure_ascii=False)
    crm_details_json = json.dumps(user_data.get('crm_details', {}), ensure_ascii=False)
    target_user_id = user_data.get('target_user_id')
    
    cursor.execute("""
        INSERT INTO orders (order_id, user_id, user_username, products_json, total_price, receipt_file_id, full_name, phone, crm_details_json, order_type, target_user_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (order_id, user.id, user.username, products_json, user_data['total_price'], user_data['receipt_file_id'],
          user_data['full_name'], user_data['phone'], crm_details_json, 'new_purchase', target_user_id))
    conn.commit()
    conn.close()
    # logging
    try:
        log_order_event(order_id, f"ORDER CREATED by @{user.username} (user_id={user.id}) products={products_json} total_price={user_data['total_price']} crm_details={crm_details_json}")
        log_order_snapshot(order_id)
    except Exception:
        logger.exception("Failed to log creation for %s", order_id)
    
    await update.message.reply_text(
        f"سفارش شما با موفقیت ثبت شد! 🎉\nشماره سفارش شما: `{order_id}`\n"
        "سفارشت در حال حاضر در انتظار تایید مدیر است. به زودی نتیجه به شما اطلاع داده می‌شود.",
        parse_mode=ParseMode.MARKDOWN
    )
    
    product_names_str = "، ".join([f"{p['name']} (تعداد: {p['quantity']})" for p in product_details])
    admin_message = (f"🔔 **سفارش جدید ثبت شد** 🔔\n\n"
                     f"شماره سفارش: `{order_id}`\n"
                     f"خریدار: {user.mention_html()} (ID: `{user.id}`)\n")
    if target_user_id:
        admin_message += f"**برای کاربر:** `{target_user_id}`\n"
    admin_message += (f"نام خریدار: {user_data['full_name']}\nتلفن خریدار: {user_data['phone']}\n\n"
                      f"**محصولات:** {product_names_str}\n"
                      f"**مبلغ:** {persian_format_number(user_data['total_price'])} تومان\n")

    if user_data.get('crm_details'):
        admin_message += "\n--- **اطلاعات اختصاصی بات CRM** ---\n"
        crm_info = user_data['crm_details']
        for key, value in crm_info.items():
            if key not in ['logo_file_id', 'product_info_file_id', 'product_info_file_name']:
                admin_message += f"▫️ **{key.replace('_', ' ').title()}:** {value}\n"
        if 'product_info_file_id' in crm_info:
            admin_message += "▫️ **اطلاعات محصولات:** در فایل زیر ارسال شده.\n"

    admin_message += "\n👇 فیش واریزی:"
    await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.HTML)
    
    if user_data.get('crm_details', {}).get('product_info_file_id'):
        file_id = user_data['crm_details']['product_info_file_id']
        file_name = user_data['crm_details']['product_info_file_name']
        await context.bot.send_document(OWNER_ID, document=file_id, filename=file_name, caption=f"فایل سفارش `{order_id}`")

    if user_data.get('crm_details', {}).get('logo_file_id') and user_data['crm_details']['logo_file_id'] != 'skipped':
        await context.bot.send_photo(OWNER_ID, user_data['crm_details']['logo_file_id'], caption="لوگوی فروشگاه مشتری")
        
    await context.bot.send_photo(OWNER_ID, user_data['receipt_file_id'])
    
    keyboard_admin = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ تایید سفارش", callback_data=f"admin_confirm_{order_id}"),
        InlineKeyboardButton("❌ رد سفارش", callback_data=f"admin_reject_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, "لطفا سفارش را تایید یا رد کنید:", reply_markup=keyboard_admin)
    
    context.user_data.clear()
    return ConversationHandler.END

# --- توابع تمدید ---
async def start_renewal(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    service_id = int(query.data.split("_")[1])
    context.user_data['renewal_service_id'] = service_id

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT product_id, product_name FROM active_services WHERE id = ?", (service_id,))
    service_info = cursor.fetchone()
    if not service_info:
        await query.edit_message_text("خطا: سرویس مورد نظر یافت نشد.")
        return MAIN_MENU
    
    product_id, product_name = service_info
    cursor.execute("SELECT price FROM products WHERE id = ?", (product_id,))
    price = cursor.fetchone()[0]
    conn.close()

    context.user_data['total_price'] = price

    if product_name == CRM_BOT_PRODUCT_NAME:
        keyboard = [
            [InlineKeyboardButton("بله، از همان اطلاعات قبلی استفاده شود", callback_data="renew_crm_yes")],
            [InlineKeyboardButton("خیر، می‌خواهم اطلاعات جدید وارد کنم", callback_data="renew_crm_no")]
        ]
        await query.edit_message_text(
            f"شما در حال تمدید **{product_name}** هستید.\nآیا می‌خواهید از همان اطلاعاتی که در خرید اول ثبت کردید استفاده شود؟",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return RENEW_CRM_CHOICE
    else:
        await query.edit_message_text(
            f"شما در حال تمدید اشتراک **{product_name}** هستید.\n"
            f"مبلغ قابل پرداخت: **{persian_format_number(price)} تومان**\n\n"
            f"لطفاً مبلغ را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
            parse_mode=ParseMode.MARKDOWN
        )
        return WAITING_FOR_RENEWAL_RECEIPT

async def renew_crm_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data

    if choice == "renew_crm_yes":
        price = context.user_data.get('total_price', 0)
        await query.edit_message_text(
            f"بسیار خب. لطفاً مبلغ **{persian_format_number(price)} تومان** را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
            parse_mode=ParseMode.MARKDOWN
        )
        return WAITING_FOR_RENEWAL_RECEIPT
    else: # renew_crm_no
        context.user_data['crm_details'] = {}
        await query.edit_message_text("لطفاً فرآیند ثبت اطلاعات جدید را از ابتدا طی کنید.\n\nابتدا **نام و نام خانوادگی** خود را وارد کنید:", parse_mode=ParseMode.MARKDOWN)
        return ASK_CRM_LOGO

async def receive_renewal_receipt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not update.message.photo:
        await update.message.reply_text("لطفا فقط عکس فیش واریزی رو بفرست.")
        return WAITING_FOR_RENEWAL_RECEIPT
    
    receipt_file_id = update.message.photo[-1].file_id
    user_data = context.user_data
    user = update.effective_user
    order_id = generate_order_id()
    service_id = user_data['renewal_service_id']
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT product_id, product_name FROM active_services WHERE id = ?", (service_id,))
    product_id, product_name = cursor.fetchone()
    
    products_json = json.dumps([{"id": product_id, "name": product_name, "price": user_data['total_price'], "quantity": 1}], ensure_ascii=False)
    crm_details_json = json.dumps(user_data.get('crm_details', {}), ensure_ascii=False)

    cursor.execute(
        "SELECT full_name, phone FROM orders WHERE user_id = ? ORDER BY timestamp DESC LIMIT 1",
        (user.id,)
    )
    last_order_info = cursor.fetchone()
    full_name = last_order_info[0] if last_order_info else "کاربر تمدیدی"
    phone = last_order_info[1] if last_order_info else "0"

    cursor.execute("""
        INSERT INTO orders (order_id, user_id, user_username, products_json, total_price, receipt_file_id, full_name, phone, crm_details_json, order_type, related_service_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (order_id, user.id, user.username, products_json, user_data['total_price'], receipt_file_id,
          full_name, phone, crm_details_json, 'renewal', service_id))
    conn.commit()
    conn.close()

    await update.message.reply_text(
        f"درخواست تمدید شما با موفقیت ثبت شد! 🎉\nشماره پیگیری: `{order_id}`\n"
        "پس از تایید مدیر، اشتراک شما تمدید خواهد شد.",
        parse_mode=ParseMode.MARKDOWN
    )
    
    admin_message = (f"🔔 **درخواست تمدید ثبت شد** 🔔\n\n"
                     f"شماره پیگیری: `{order_id}`\n"
                     f"کاربر: {user.mention_html()} (ID: `{user.id}`)\n"
                     f"سرویس برای تمدید: **{product_name}**\n"
                     f"مبلغ: {persian_format_number(user_data['total_price'])} تومان\n\n"
                     f"👇 فیش واریزی:")
    await context.bot.send_message(OWNER_ID, admin_message, parse_mode=ParseMode.HTML)
    await context.bot.send_photo(OWNER_ID, receipt_file_id)
    
    keyboard_admin = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ تایید تمدید", callback_data=f"admin_confirm_{order_id}"),
        InlineKeyboardButton("❌ رد تمدید", callback_data=f"admin_reject_{order_id}")
    ]])
    await context.bot.send_message(OWNER_ID, "لطفا درخواست تمدید را تایید یا رد کنید:", reply_markup=keyboard_admin)
    
    context.user_data.clear()
    return ConversationHandler.END

# --- پنل ادمین ---
async def admin_pannel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID:
        await update.message.reply_text("شما اجازه دسترسی به این بخش را ندارید.")
        return
    reply_markup = admin_pannel_keyboard()
    if update.callback_query:
        try:
            await update.callback_query.edit_message_text("👑 به پنل مدیریت خوش آمدید. لطفاً بخش مورد نظر را انتخاب کنید:", reply_markup=reply_markup)
        except BadRequest: # Message is not modified
            pass
    else:
        await update.message.reply_text("👑 به پنل مدیریت خوش آمدید. لطفاً بخش مورد نظر را انتخاب کنید:", reply_markup=reply_markup)

def admin_pannel_keyboard():
    keyboard = [
        [InlineKeyboardButton("⏳ سفارشات و تمدیدهای در انتظار", callback_data="admin_view_pending_approval")],
        [InlineKeyboardButton("✨ درخواست‌های کاستوم بات AI", callback_data="admin_view_pending_admin_approval")],
        [InlineKeyboardButton("💬 درخواست‌های قیمت‌گذاری چت‌بات", callback_data="admin_view_pending_quote")],
        [InlineKeyboardButton("👥 مدیریت کاربران", callback_data="admin_view_users_0")],
        [InlineKeyboardButton("✅ سفارشات تایید شده", callback_data="admin_view_confirmed")],
        [InlineKeyboardButton("❌ سفارشات رد شده", callback_data="admin_view_rejected")],
        [InlineKeyboardButton("🚚 سفارشات تکمیل شده", callback_data="admin_view_completed")],
        [InlineKeyboardButton("📦 مشاهده همه سفارشات", callback_data="admin_view_all")],
    ]
    return InlineKeyboardMarkup(keyboard)


async def _send_chunks(bot, chat_id: int, text: str, parse_mode=ParseMode.MARKDOWN):
    """Send a long text in chunks to avoid Telegram limits."""
    max_len = 3800
    cur = ""
    for line in text.split("\n"):
        if len(cur) + len(line) + 1 > max_len:
            await bot.send_message(chat_id, cur, parse_mode=parse_mode)
            cur = line + "\n"
        else:
            cur += line + "\n"
    if cur:
        await bot.send_message(chat_id, cur, parse_mode=parse_mode)


async def _run_db(func, *args, **kwargs):
    """Run a blocking DB function in the default thread pool."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, partial(func, *args, **kwargs))


async def _admin_list_users(query, context, page: int = 0, per_page: int = 12):
    offset = page * per_page

    def _fetch():
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT COUNT(DISTINCT user_id) FROM (SELECT user_id FROM orders UNION SELECT user_id FROM active_services)")
            total_users = cursor.fetchone()[0] or 0
        except Exception:
            total_users = 0

        cursor.execute(
            """
            SELECT u.user_id,
                   MAX(o.user_username) as username,
                   MAX(o.full_name) as full_name,
                   MAX(o.timestamp) as last_ts
            FROM (SELECT user_id FROM orders UNION SELECT user_id FROM active_services) AS u
            LEFT JOIN orders o ON u.user_id = o.user_id
            GROUP BY u.user_id
            ORDER BY last_ts DESC
            LIMIT ? OFFSET ?
            """,
            (per_page, offset)
        )
        users = cursor.fetchall()
        conn.close()
        return total_users, users

    total_users, users = await _run_db(_fetch)

    back_button = InlineKeyboardButton("🔙 بازگشت به پنل اصلی", callback_data="admin_back_to_panel")
    if not users:
        try:
            await query.edit_message_text("هیچ کاربری یافت نشد.", reply_markup=InlineKeyboardMarkup([[back_button]]))
        except Exception:
            await context.bot.send_message(OWNER_ID, "هیچ کاربری یافت نشد.")
        return

    message_text = f"📋 لیست مشتریان (نمایش {offset+1} - {offset+len(users)} از {total_users}):\n\n"
    keyboard = []
    for user_row in users:
        user_id, username, full_name, _ = user_row
        label = f"{user_id}"
        if username:
            label = f"@{username} — {user_id}"
        elif full_name:
            label = f"{full_name} — {user_id}"
        keyboard.append([InlineKeyboardButton(label, callback_data=f"admin_view_user_{user_id}")])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ قبلی", callback_data=f"admin_view_users_{page-1}"))
    if (page + 1) * per_page < total_users:
        nav.append(InlineKeyboardButton("بعدی ➡️", callback_data=f"admin_view_users_{page+1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([back_button])

    try:
        await query.edit_message_text(message_text, reply_markup=InlineKeyboardMarkup(keyboard))
    except BadRequest:
        await context.bot.send_message(OWNER_ID, message_text, reply_markup=InlineKeyboardMarkup(keyboard))


async def _admin_show_user(query, context, user_id: int):
    def _fetch():
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT user_username, full_name, phone FROM orders WHERE user_id = ? ORDER BY timestamp DESC LIMIT 1", (user_id,))
        prof = cursor.fetchone() or (None, None, None)

        cursor.execute("SELECT product_name, expiry_date, is_active FROM active_services WHERE user_id = ? ORDER BY expiry_date DESC", (user_id,))
        services = cursor.fetchall()

        cursor.execute("""SELECT order_id, status, total_price, timestamp, products_json, order_type, crm_details_json, receipt_file_id
                          FROM orders WHERE user_id = ? ORDER BY timestamp DESC""", (user_id,))
        orders = cursor.fetchall()
        conn.close()
        return prof, services, orders

    prof, services, orders = await _run_db(_fetch)

    username, full_name, phone = prof
    username = username or ""
    full_name = full_name or ""
    phone = phone or ""

    lines = [
        f"👤 پروفایل کاربر — ID: `{user_id}`",
        f"- یوزرنیم: @{username if username else 'N/A'}",
        f"- نام ثبت‌شده: {full_name or 'N/A'}",
        f"- تلفن (آخرین ثبت‌شده): {phone or 'N/A'}",
        f"- تعداد سفارش‌ها: {len(orders)}",
        "",
        "📦 سرویس‌های فعال/غیرفعال:"
    ]

    if services:
        today = datetime.now().date()
        for pname, expiry, is_active in services:
            active_text = "فعال" if is_active == 1 or is_active is True else "غیرفعال"
            if not expiry or expiry in ("", "دائمی"):
                lines.append(f"- {pname} ({active_text})")
            else:
                try:
                    expd = datetime.strptime(expiry, "%Y-%m-%d").date()
                    rem = (expd - today).days
                    if rem < 0: rem = 0
                    lines.append(f"- {pname} — {rem} روز مانده ({active_text}) (انقضا: {expiry})")
                except Exception:
                    lines.append(f"- {pname} — تاریخ نامعتبر: {expiry} ({active_text})")
    else:
        lines.append("- ندارد")

    lines.append("\n🧾 سفارش‌ها (برای مشاهده جزئیات، روی هر سفارش کلیک کنید):")
    keyboard = []
    if orders:
        # small helpers local to this view
        def _shorten(text, n=80):
            try:
                text = str(text)
            except Exception:
                return ""
            return text if len(text) <= n else text[: n - 1] + "…"

        def _extract_descs(crm_str):
            user_desc = None
            admin_desc = None
            if not crm_str:
                return user_desc, admin_desc
            try:
                parsed = json.loads(crm_str)
            except Exception:
                # plain text stored as user description
                return crm_str, None

            if isinstance(parsed, dict):
                user_desc = parsed.get('description') or parsed.get('user_description') or parsed.get('desc')
                admin_desc = parsed.get('details') or parsed.get('admin_details')
            else:
                # non-dict JSON (e.g., list) -> show as user desc
                user_desc = str(parsed)
            return user_desc, admin_desc

        for o in orders:
            oid, status, total_price, ts, pj, order_type, crm_details_json, receipt_file_id = o
            prod_name = "نامشخص"
            try:
                products = json.loads(pj) if pj else []
                if isinstance(products, list) and products:
                    first = products[0]
                    prod_name = first.get('name') if isinstance(first, dict) else str(first)
            except Exception:
                pass
            price_text = f"{persian_format_number(total_price)} تومان" if isinstance(total_price, (int, float)) else str(total_price)
            lines.append(f"• `{oid}` | {prod_name} | وضعیت: {status} | مبلغ: {price_text}")

            # show short user/admin descriptions (if present)
            try:
                user_d, admin_d = _extract_descs(crm_details_json)
                if user_d:
                    lines.append(f"    - توضیحات کاربر: {_shorten(user_d, 120)}")
                if admin_d:
                    lines.append(f"    - توضیحات ادمین: {_shorten(admin_d, 120)}")
            except Exception:
                pass

            keyboard.append([InlineKeyboardButton(f"{prod_name} — {oid}", callback_data=f"admin_view_order_{oid}")])
    else:
        lines.append("- کاربر هنوز سفارشی ثبت نکرده است.")

    keyboard.append([InlineKeyboardButton("⬅️ بازگشت به لیست مشتریان", callback_data="admin_view_users_0")])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل اصلی", callback_data="admin_back_to_panel")])

    message_text = "\n".join(lines)
    try:
        await query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
    except BadRequest:
        # fallback: send chunks to admin
        try:
            await query.edit_message_text("در حال ارسال اطلاعات کامل کاربر ...")
        except Exception:
            pass
        await _send_chunks(context.bot, OWNER_ID, message_text)
        try:
            await context.bot.send_message(OWNER_ID, "عملیات:", reply_markup=InlineKeyboardMarkup(keyboard))
        except Exception:
            pass


async def _admin_show_order(query, context, order_id: str):
    # Clean, defensive implementation
    try:
        def _fetch():
            conn = sqlite3.connect("shop_data.db", check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute("""SELECT order_id, user_id, user_username, products_json, total_price, status, receipt_file_id, crm_details_json, full_name, phone, timestamp, order_type
                              FROM orders WHERE order_id = ?""", (order_id,))
            order = cursor.fetchone()
            conn.close()
            return order

        order = await _run_db(_fetch)
        if not order:
            try:
                await query.edit_message_text("سفارش مورد نظر یافت نشد.")
            except Exception:
                pass
            return

        (oid, uid, uusername, products_json_str, total_price, status, receipt_file_id,
         crm_details_str, ofull, ophone, ts, order_type) = order

        lines = [
            f"🧾 جزئیات سفارش `{oid}`",
            f"- کاربر: @{uusername if uusername else 'N/A'} (ID: `{uid}`)",
            f"- نام ثبت‌شده: {ofull or 'N/A'}",
            f"- تلفن ثبت‌شده: {ophone or 'N/A'}",
            f"- وضعیت: {status}",
            f"- نوع سفارش: {order_type}",
            f"- مبلغ: {persian_format_number(total_price) if isinstance(total_price, (int, float)) else total_price}",
            f"- زمان ثبت: {ts}",
            "",
            "🔹 محصولات/خدمات:"
        ]

        # products
        try:
            products = json.loads(products_json_str) if products_json_str else []
            if isinstance(products, list) and products:
                for p in products:
                    if isinstance(p, dict):
                        pname = p.get('name', 'نامشخص')
                        qty = p.get('quantity') or p.get('qty') or 1
                        pprice = p.get('price')
                        if pprice is not None:
                            lines.append(f"• {pname} x{qty} — {persian_format_number(pprice)} تومان")
                        else:
                            lines.append(f"• {pname} x{qty}")
                    else:
                        lines.append(f"• {str(p)}")
            else:
                lines.append("— اطلاعات محصول ناموجود")
        except Exception:
            lines.append("— خطا در پارس کردن محصولات")

        # crm details: try JSON first, fall back to raw text
        if crm_details_str:
            lines.append("\n🔸 جزئیات تکمیلی سفارش:")
            try:
                details = None
                try:
                    details = json.loads(crm_details_str)
                except Exception:
                    # treat as plain user-provided text
                    lines.append(f"- توضیحات کاربر: {crm_details_str}")

                if isinstance(details, dict):
                    user_desc = details.get('description') or details.get('user_description') or details.get('desc')
                    if user_desc:
                        lines.append(f"- توضیحات کاربر: {user_desc}")

                    admin_details = details.get('details') or details.get('admin_details')
                    price = details.get('price') or details.get('quote_price') or details.get('amount')
                    admin_parts = []
                    if price is not None:
                        try:
                            admin_parts.append(f"قیمت: {persian_format_number(int(price))} تومان")
                        except Exception:
                            admin_parts.append(f"قیمت: {price}")
                    if admin_details:
                        admin_parts.append(str(admin_details))
                    if admin_parts:
                        lines.append(f"- توضیحات ادمین: {' — '.join(admin_parts)}")

                    for k in ('type', 'full_name', 'phone'):
                        if k in details and details.get(k):
                            lines.append(f"- {k}: {details.get(k)}")

                    other_keys = [kk for kk in details.keys() if kk not in ('description', 'user_description', 'desc', 'details', 'admin_details', 'price', 'quote_price', 'amount', 'type', 'full_name', 'phone')]
                    if other_keys:
                        lines.append(f"- سایر اطلاعات: {', '.join(other_keys)}")
            except Exception:
                logger.exception("Failed to parse crm_details_json for order %s", oid)
                lines.append("- (جزئیات قابل نمایش نیست)")

        order_text = "\n".join(lines)
        try:
            await query.edit_message_text(order_text, parse_mode=ParseMode.MARKDOWN)
        except BadRequest:
            await context.bot.send_message(OWNER_ID, order_text, parse_mode=ParseMode.MARKDOWN)

        # send attached files
        try:
            if crm_details_str:
                try:
                    details = json.loads(crm_details_str)
                except Exception:
                    details = {}
                doc_id = details.get('file_id') or details.get('product_info_file_id')
                if doc_id:
                    try:
                        await context.bot.send_document(OWNER_ID, doc_id, caption=f"فایل سفارش `{oid}`")
                    except Exception as e:
                        logger.error("Failed to send order attached document: %s", e)
                logo_id = details.get('logo_file_id')
                if logo_id and logo_id != 'skipped':
                    try:
                        await context.bot.send_photo(OWNER_ID, logo_id, caption="لوگوی مشتری")
                    except Exception as e:
                        logger.error("Failed to send order logo: %s", e)
            if receipt_file_id:
                try:
                    await context.bot.send_photo(OWNER_ID, receipt_file_id, caption="فیش واریزی")
                except Exception:
                    try:
                        await context.bot.send_document(OWNER_ID, receipt_file_id, caption="فیش واریزی")
                    except Exception as e:
                        logger.error("Failed to send receipt for order %s: %s", oid, e)
        except Exception as e:
            logger.error("Error while sending files for order %s: %s", oid, e)

        # action buttons after details
        buttons = []
        if status == "pending_approval":
            buttons.append([InlineKeyboardButton("✅ تایید سفارش", callback_data=f"admin_confirm_{oid}"),
                            InlineKeyboardButton("❌ رد سفارش", callback_data=f"admin_reject_{oid}")])
        if status == "pending_quote":
            buttons.append([InlineKeyboardButton("💬 قیمت‌گذاری", callback_data=f"admin_quote_{oid}")])
        if status == "pending_admin_approval":
            buttons.append([InlineKeyboardButton("✅ تایید و قیمت‌گذاری", callback_data=f"admin_approve_custom_{oid}"),
                            InlineKeyboardButton("❌ رد درخواست", callback_data=f"admin_reject_custom_{oid}")])
        if status == "confirmed":
            buttons.append([InlineKeyboardButton("🚚 اعلام تکمیل", callback_data=f"admin_complete_{oid}")])

        buttons.append([InlineKeyboardButton("🔙 بازگشت به کاربر", callback_data=f"admin_view_user_{uid}")])
        buttons.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data="admin_back_to_panel")])

        try:
            await context.bot.send_message(OWNER_ID, "اقدامات مربوط به سفارش:", reply_markup=InlineKeyboardMarkup(buttons))
        except Exception:
            pass

    except Exception as e:
        logger.exception("Unhandled error in _admin_show_order for %s: %s", order_id, e)
        # Provide a safe back keyboard so admin can easily return
        back_buttons = []
        try:
            # if uid available in scope, offer back to that user
            if 'uid' in locals() and uid:
                back_buttons.append(InlineKeyboardButton("🔙 بازگشت به کاربر", callback_data=f"admin_view_user_{uid}"))
        except Exception:
            pass
        back_buttons.append(InlineKeyboardButton("🔙 بازگشت به پنل", callback_data="admin_back_to_panel"))
        kb = InlineKeyboardMarkup([back_buttons])
        try:
            await query.edit_message_text("خطا در بارگذاری جزئیات سفارش. لطفاً لاگ‌ها را بررسی کنید.", reply_markup=kb)
        except Exception:
            try:
                await context.bot.send_message(OWNER_ID, "خطا در بارگذاری جزئیات سفارش. لطفاً لاگ‌ها را بررسی کنید.", reply_markup=kb)
            except Exception:
                pass
        return

async def admin_actions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    data = query.data.split("_")
    action_type = data[1]
    
    # --- START: منطق جدید برای کاستوم بات ---
    if action_type == "approve" and data[2] == "custom":
        order_id = "_".join(data[3:])
        context.user_data['next_admin_action'] = 'get_custom_ai_price'
        context.user_data['target_order_id'] = order_id
        await query.edit_message_text(f"در حال قیمت‌گذاری برای کاستوم بات `{order_id}`.\n\nلطفاً **مبلغ کل** را به تومان (فقط عدد) وارد کنید:", parse_mode=ParseMode.MARKDOWN)
        return

    if action_type == "reject" and data[2] == "custom":
        order_id = "_".join(data[3:])
        context.user_data['next_admin_action'] = 'get_custom_ai_reject_reason'
        context.user_data['target_order_id'] = order_id
        await query.edit_message_text(f"در حال رد کردن درخواست `{order_id}`.\n\nلطفاً **دلیل رد کردن** را در یک پیام وارد کنید:", parse_mode=ParseMode.MARKDOWN)
        return
    # --- END: منطق جدید برای کاستوم بات ---

    if action_type == "back" and data[2] == "to" and data[3] == "panel":
        await admin_pannel(update, context)
        return
    
    if action_type == "quote":
        order_id = "_".join(data[2:])
        context.user_data['next_admin_action'] = 'get_quote_price'
        context.user_data['target_order_id'] = order_id
        await query.edit_message_text(f"در حال قیمت‌گذاری برای سفارش `{order_id}`.\n\nلطفاً **مبلغ کل** را به تومان (فقط عدد) وارد کنید:", parse_mode=ParseMode.MARKDOWN)
        return

    if action_type in ["confirm", "reject", "complete"]:
        order_id = "_".join(data[2:])
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, status, products_json, order_type, related_service_id, full_name, phone, target_user_id FROM orders WHERE order_id = ?", (order_id,))
        result = cursor.fetchone()
        
        if not result:
            await query.edit_message_text("خطا: سفارش یافت نشد.")
            conn.close()
            return
        
        user_id, current_status, products_json_str, order_type, related_service_id, full_name, phone, target_user_id = result
        products = json.loads(products_json_str)
        
        new_status, user_message, admin_feedback = "", "", ""

        if action_type == "confirm" and current_status == 'pending_approval':
            if order_type in ['ai_credit', 'chatbot_request', 'custom_ai_request'] or (order_type == 'new_purchase' and any(p['name'] == CRM_BOT_PRODUCT_NAME for p in products)):
                context.user_data['next_admin_action'] = 'get_bot_username'
                context.user_data['target_order_id'] = order_id
                prompt_text = "این سفارش شامل بات اختصاصی است.\nلطفاً **نام کاربری ربات** ساخته شده برای مشتری را ارسال کنید (مثلاً @CustomerBot):"
                await query.edit_message_text(prompt_text, parse_mode=ParseMode.MARKDOWN)
                conn.close()
                return
            
            new_status = "confirmed"
            admin_feedback = f"✅ سفارش/تمدید {order_id} تایید شد."
            
            if order_type == 'new_purchase':
                expiry_date = (datetime.now() + timedelta(days=30)).date().isoformat()
                product_names = []
                
                service_user_id = target_user_id if target_user_id else user_id

                for p in products:
                    cursor.execute(
                        "INSERT INTO active_services (user_id, product_id, product_name, expiry_date) VALUES (?, ?, ?, ?)",
                        (service_user_id, p['id'], p['name'], expiry_date)
                    )
                    product_names.append(p['name'])
                
                product_names_str = "، ".join(product_names)
                user_message = f"خبر خوب! 😍\nسفارش شما برای سرویس(های) **{product_names_str}** با شماره `{order_id}` تایید شد و اشتراک(های) شما فعال گردید."
                
                extra_info = []
                if "بات تریدر" in product_names: 
                    extra_info.append(f"نام کاربری ربات تریدر: {TRADER_BOT_USERNAME}")
                
                if extra_info:
                    user_message += "\n\n" + "\n".join(extra_info)
                
                log_to_excel(full_name, phone, product_names_str, order_id, "تایید شده", 30, target_user_id)

            elif order_type == 'renewal':
                product_name = products[0]['name']
                cursor.execute("SELECT expiry_date FROM active_services WHERE id = ?", (related_service_id,))
                current_expiry_str = cursor.fetchone()[0]
                current_expiry = datetime.strptime(current_expiry_str, '%Y-%m-%d').date()
                if current_expiry < datetime.now().date():
                    new_expiry_date = (datetime.now() + timedelta(days=30)).date()
                else:
                    new_expiry_date = (current_expiry + timedelta(days=30))
                
                cursor.execute(
                    "UPDATE active_services SET expiry_date = ?, is_active = 1 WHERE id = ?",
                    (new_expiry_date.isoformat(), related_service_id)
                )
                remaining_days = (new_expiry_date - datetime.now().date()).days
                user_message = f"خبر خوب! 😍\nتمدید اشتراک شما برای سرویس **{product_name}** با موفقیت انجام شد."
                log_to_excel(full_name, phone, product_name, order_id, "تمدید شده", remaining_days)

        elif action_type == "reject" and current_status in ['pending_approval', 'confirmed', 'user_approved']:
            new_status = "rejected"
            if current_status == 'confirmed':
                if order_type == 'new_purchase':
                    for p in products:
                        cursor.execute("DELETE FROM active_services WHERE user_id = ? AND product_id = ?", (user_id, p['id']))
            user_message = (f"متاسفانه درخواست شما با شماره پیگیری `{order_id}` رد شد. 😔\n"
                            f"برای پیگیری، لطفاً با پشتیبانی تماس بگیرید:\n{SUPPORT_PHONE}")
            admin_feedback = f"❌ درخواست {order_id} رد شد."
            update_excel_status(order_id, "رد شده")

        elif action_type == "complete" and current_status == 'confirmed':
            new_status = "completed"
            product_names_str = "، ".join([p['name'] for p in products])
            user_message = f"سفارش شما با شماره `{order_id}` برای سرویس(های) **{product_names_str}** تکمیل شد. از خرید شما متشکریم!"
            admin_feedback = f"✅ سفارش {order_id} به وضعیت 'تکمیل شده' تغییر یافت."
            update_excel_status(order_id, "تکمیل شده")
        
        if new_status:
            cursor.execute("UPDATE orders SET status = ? WHERE order_id = ?", (new_status, order_id))
            conn.commit()
            if user_message:
                try: await context.bot.send_message(chat_id=user_id, text=user_message, parse_mode=ParseMode.MARKDOWN)
                except Exception as e:
                    logger.error(f"Failed to send message to user {user_id}: {e}")
                    admin_feedback += f"\n⚠️ **اخطار:** ارسال پیام به کاربر {user_id} ناموفق بود."
        
        conn.close()
        await query.edit_message_text(admin_feedback)

# ...existing code...
    elif action_type == "view":
        # delegate view actions to helper functions
        try:
            if data[2] == "users":
                try:
                    page = int(data[3]) if len(data) > 3 else 0
                except Exception:
                    page = 0
                await _admin_list_users(query, context, page=page, per_page=12)
                return

            if data[2] == "user":
                # callback format: admin_view_user_<user_id>
                user_id = int("_".join(data[3:]))
                await _admin_show_user(query, context, user_id)
                return

            if data[2] == "order":
                order_id = "_".join(data[3:])
                await _admin_show_order(query, context, order_id)
                return

            if data[2] == "user":
                # show full details for a specific user (تمام سفارش‌ها، شماره تماس و اطلاعات)
                try:
                    user_id = int(data[3])
                except Exception:
                    await query.edit_message_text("شناسه کاربر نامعتبر است.")
                    return

                conn = sqlite3.connect("shop_data.db", check_same_thread=False)
                cursor = conn.cursor()

                # latest profile info (from last order if available)
                cursor.execute("SELECT full_name, user_username, phone FROM orders WHERE user_id = ? ORDER BY timestamp DESC LIMIT 1", (user_id,))
                profile = cursor.fetchone() or (None, None, None)
                full_name, username, phone = profile

                # active services
                cursor.execute("SELECT product_name, expiry_date FROM active_services WHERE user_id = ? AND is_active = 1", (user_id,))
                services = cursor.fetchall()

                # all orders for this user (show all as requested)
                cursor.execute("SELECT order_id, status, total_price, timestamp, products_json, order_type, crm_details_json FROM orders WHERE user_id = ? ORDER BY timestamp DESC", (user_id,))
                user_orders = cursor.fetchall()

                conn.close()

                message_lines = []
                message_lines.append(f"👤 اطلاعات کامل کاربر (ID: `{user_id}`):")
                if full_name: message_lines.append(f"- نام ثبت‌شده: {full_name}")
                if username: message_lines.append(f"- یوزرنیم: @{username}")
                if phone: message_lines.append(f"- تلفن: {phone}")
                message_lines.append(f"- تعداد سفارش‌ها: {len(user_orders)}")
                message_lines.append("\n📦 سرویس‌های فعال:")
                if services:
                    today = datetime.now().date()
                    for svc_name, expiry in services:
                        if expiry in (None, "", "دائمی"):
                            message_lines.append(f"- {svc_name} (دائمی/اعتباری)")
                        else:
                            try:
                                expiry_date = datetime.strptime(expiry, "%Y-%m-%d").date()
                                remaining = (expiry_date - today).days
                                if remaining < 0: remaining = 0
                                message_lines.append(f"- {svc_name} — {remaining} روز مانده")
                            except Exception:
                                message_lines.append(f"- {svc_name} — تاریخ: {expiry}")
                else:
                    message_lines.append("- هیچ سرویس فعالی ندارد.")

                message_lines.append("\n🧾 سفارش‌ها (جدیدترین در بالا):")
                if user_orders:
                    for o in user_orders:
                        try:
                            # SELECT order_id, status, total_price, timestamp, products_json, order_type, crm_details_json, full_name, phone
                            oid, status, total_price, ts, pj, order_type, crm_details_str, o_full_name, o_phone = o
                        except Exception:
                            # try a shorter unpack if DB doesn't include full_name/phone per-order
                            try:
                                oid, status, total_price, ts, pj, order_type, crm_details_str = o
                                o_full_name, o_phone = None, None
                            except Exception:
                                continue

                        # prefer per-order name/phone if available, else use profile top-level
                        display_name = o_full_name or full_name or "N/A"
                        display_phone = o_phone or phone or "N/A"

                        # parse محصولات
                        products_desc = []
                        try:
                            products = json.loads(pj) if pj else []
                            if isinstance(products, list) and products:
                                for p in products:
                                    if isinstance(p, dict):
                                        pname = p.get('name', 'نامشخص')
                                        qty = p.get('quantity') or p.get('qty') or 1
                                        pprice = p.get('price') or p.get('unit_price')
                                        if pprice is not None:
                                            products_desc.append(f"{pname} x{qty} ({persian_format_number(pprice)} تومان)")
                                        else:
                                            products_desc.append(f"{pname} x{qty}")
                            elif products:
                                products_desc.append(str(products))
                        except Exception:
                            products_desc.append("اطلاعات محصول نامعتبر")

                        prod_text = "؛ ".join(products_desc) if products_desc else "—"
                        price_text = f"{persian_format_number(total_price)} تومان" if isinstance(total_price, (int, float)) and total_price >= 0 else str(total_price)
                        ts_text = ts if ts else "N/A"

                        message_lines.append(f"• `{oid}` | وضعیت: {status} | مبلغ پرداختی: {price_text} | زمان: {ts_text}")
                        message_lines.append(f"   - سرویس(ها): {prod_text}")
                        message_lines.append(f"   - نام (ثبت‌شده): {display_name} — تلفن: {display_phone}")

                        # crm/custom details summary: show user description + admin notes/price
                        if crm_details_str:
                            try:
                                details_obj = json.loads(crm_details_str)
                                user_desc = details_obj.get('description') or details_obj.get('user_description') or details_obj.get('desc')
                                if user_desc:
                                    message_lines.append(f"   - توضیحات کاربر: {user_desc}")

                                admin_details = details_obj.get('details') or details_obj.get('admin_details')
                                price = details_obj.get('price') or details_obj.get('quote_price') or details_obj.get('amount')
                                admin_parts = []
                                if price is not None:
                                    try:
                                        admin_parts.append(f"قیمت: {persian_format_number(int(price))} تومان")
                                    except Exception:
                                        admin_parts.append(f"قیمت: {price}")
                                if admin_details:
                                    admin_parts.append(str(admin_details))
                                if admin_parts:
                                    message_lines.append(f"   - توضیحات ادمین: {' — '.join(admin_parts)}")
                            except Exception:
                                pass
                        message_lines.append("")
                else:
                    message_lines.append("- کاربر هیچ سفارشی ثبت نکرده است.")

                message_text = "\n".join(message_lines)

                # build per-order admin action buttons
                keyboard_rows = []
                for oid, status, total_price, ts, pj, order_type, crm_details_str in user_orders:
                    if status == "pending_approval":
                        keyboard_rows.append([
                            InlineKeyboardButton(f"✅ تایید: {oid}", callback_data=f"admin_confirm_{oid}"),
                            InlineKeyboardButton(f"❌ رد: {oid}", callback_data=f"admin_reject_{oid}")
                        ])
                    elif status == "pending_quote":
                        keyboard_rows.append([InlineKeyboardButton(f"💬 قیمت‌گذاری: {oid}", callback_data=f"admin_quote_{oid}")])
                    elif status == "pending_admin_approval":
                        keyboard_rows.append([
                            InlineKeyboardButton(f"✅ تایید/قیمت: {oid}", callback_data=f"admin_approve_custom_{oid}"),
                            InlineKeyboardButton(f"❌ رد: {oid}", callback_data=f"admin_reject_custom_{oid}")
                        ])
                    elif status == "confirmed":
                        keyboard_rows.append([
                            InlineKeyboardButton(f"🚚 تکمیل: {oid}", callback_data=f"admin_complete_{oid}"),
                            InlineKeyboardButton(f"❌ رد: {oid}", callback_data=f"admin_reject_{oid}")
                        ])

                # navigation
                keyboard_rows.append([InlineKeyboardButton("⬅️ بازگشت به لیست مشتریان", callback_data="admin_view_users_0")])
                keyboard_rows.append([InlineKeyboardButton("🔙 بازگشت به پنل اصلی", callback_data="admin_back_to_panel")])

                reply_markup = InlineKeyboardMarkup(keyboard_rows)

                # Try to edit message; if too long or edit fails, fallback to sending as separate messages
                try:
                    await query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
                except BadRequest as e:
                    logger.warning("admin_actions:user view edit failed: %s. Falling back to send_message.", e)
                    try:
                        await query.edit_message_text("در حال ارسال اطلاعات کامل کاربر به صورت پیام‌های جداگانه... لطفاً پیام‌ها را بررسی کنید.")
                    except Exception:
                        pass

                    # split into chunks under Telegram limit (~4000). use safe margin.
                    max_len = 3800
                    lines = message_text.split("\n")
                    chunks = []
                    cur = ""
                    for line in lines:
                        if len(cur) + len(line) + 1 > max_len:
                            chunks.append(cur)
                            cur = line + "\n"
                        else:
                            cur += line + "\n"
                    if cur:
                        chunks.append(cur)

                    # send chunks to admin
                    for chunk in chunks:
                        try:
                            await context.bot.send_message(OWNER_ID, chunk, parse_mode=ParseMode.MARKDOWN)
                        except Exception as send_err:
                            logger.error("Failed to send user-detail chunk to admin: %s", send_err)

                    # send navigation keyboard separately
                    try:
                        await context.bot.send_message(OWNER_ID, "عملیات:", reply_markup=reply_markup)
                    except Exception as km_err:
                        logger.error("Failed to send admin navigation keyboard: %s", km_err)

                return

            status_map = {
                "pending_approval": "⏳ در انتظار پرداخت", "confirmed": "✅ تایید شده",
                "rejected": "❌ رد شده", "completed": "🚚 تکمیل شده", "all": "همه",
                "pending_quote": "💬 در انتظار قیمت‌گذاری چت‌بات",
                "pending_admin_approval": "✨ در انتظار بررسی ادمین (کاستوم AI)" # <<<< وضعیت جدید
            }
            status_to_query = "_".join(data[2:])
            status_persian = status_map.get(status_to_query, status_to_query)

            conn = sqlite3.connect("shop_data.db", check_same_thread=False)
            cursor = conn.cursor()
            if status_to_query == "all":
                cursor.execute("SELECT order_id, user_username, total_price, status, order_type, products_json FROM orders ORDER BY timestamp DESC")
            else:
                cursor.execute("SELECT order_id, user_username, total_price, status, order_type, products_json FROM orders WHERE status = ? ORDER BY timestamp DESC", (status_to_query,))
            orders = cursor.fetchall()

            back_button = InlineKeyboardButton("🔙 بازگشت به پنل اصلی", callback_data="admin_back_to_panel")
            if not orders:
                await query.edit_message_text(f"هیچ سفارشی با وضعیت '{status_persian}' یافت نشد.", reply_markup=InlineKeyboardMarkup([[back_button]]))
                conn.close()
                return

            message_text = f"لیست سفارشات **{status_persian}**:\n\n"
            keyboard_list = []
            for order_row in orders:
                # defensive unpack (in case DB row shape changed)
                try:
                    order_id, username, total_price, status, order_type, products_json_str = order_row
                except Exception:
                    # fallback: skip malformed row
                    logger.error("Malformed order row in admin view: %s", order_row)
                    continue

                # safe parse product name
                product_info = "نامشخص"
                try:
                    products = json.loads(products_json_str) if products_json_str else []
                    if isinstance(products, list) and products:
                        # product may be dict or simple name string
                        first = products[0]
                        if isinstance(first, dict):
                            product_info = first.get('name', 'نامشخص')
                        else:
                            product_info = str(first)
                except Exception:
                    logger.exception("Failed to parse products_json for order %s", order_id)

                username_display = username or "N/A"
                message_text += (f"🔹 **{product_info}** (@{username_display})\n"
                                 f"   - شماره: `{order_id}`\n")
                if isinstance(total_price, (int, float)) and total_price > 0:
                    message_text += f"   - مبلغ: {persian_format_number(total_price)} تومان\n\n"
                else:
                    message_text += "\n"

                # buttons per-order based on status
                if status == "pending_approval":
                    keyboard_list.append([
                        InlineKeyboardButton(f"✅ تایید: {order_id}", callback_data=f"admin_confirm_{order_id}"),
                        InlineKeyboardButton(f"❌ رد: {order_id}", callback_data=f"admin_reject_{order_id}")
                    ])
                elif status == "pending_quote":
                    keyboard_list.append([
                        InlineKeyboardButton(f"✅ قیمت‌گذاری: {order_id}", callback_data=f"admin_quote_{order_id}")
                    ])
                elif status == "pending_admin_approval":
                    keyboard_list.append([
                        InlineKeyboardButton(f"✅ تایید و قیمت‌گذاری: {order_id}", callback_data=f"admin_approve_custom_{order_id}"),
                        InlineKeyboardButton(f"❌ رد کردن: {order_id}", callback_data=f"admin_reject_custom_{order_id}")
                    ])
                elif status == "confirmed":
                    keyboard_list.append([
                        InlineKeyboardButton(f"🚚 تکمیل: {order_id}", callback_data=f"admin_complete_{order_id}"),
                        InlineKeyboardButton(f"❌ رد کردن: {order_id}", callback_data=f"admin_reject_{order_id}")
                    ])

            # always add back button
            keyboard_list.append([back_button])
            reply_markup = InlineKeyboardMarkup(keyboard_list)
            await query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
            conn.close()
        except Exception as e:
            logger.exception("admin_actions:view failed: %s", e)
            try:
                await query.edit_message_text("خطا در نمایش لیست سفارشات. لطفاً لاگ‌ها را بررسی کنید.")
            except Exception:
                pass
# ...existing code...
        
        keyboard_list.append(back_button)
        reply_markup = InlineKeyboardMarkup(keyboard_list)
        await query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)

async def admin_message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id != OWNER_ID: return
    action = context.user_data.get('next_admin_action')
    if not action: return

    # --- START: منطق جدید برای دریافت پیام از ادمین ---
    if action == 'get_custom_ai_reject_reason':
        reason = update.message.text
        order_id = context.user_data['target_order_id']
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM orders WHERE order_id = ?", (order_id,))
        result = cursor.fetchone()
        if result:
            user_id = result[0]
            cursor.execute("UPDATE orders SET status = 'rejected' WHERE order_id = ?", (order_id,))
            conn.commit()
            user_message = (f"متاسفانه درخواست شما برای کاستوم بات با شماره `{order_id}` رد شد. 😔\n\n"
                            f"**دلیل از طرف ادمین:**\n{reason}")
            try:
                await context.bot.send_message(user_id, user_message, parse_mode=ParseMode.MARKDOWN)
                await update.message.reply_text(f"✅ درخواست رد شد و دلیل برای کاربر ارسال گردید.")
            except Exception as e:
                logger.error(f"Failed to send rejection reason to user {user_id}: {e}")
                await update.message.reply_text("⚠️ خطا در ارسال پیام به کاربر.")
        else:
            await update.message.reply_text("خطا: سفارش یافت نشد.")
        conn.close()
        context.user_data.clear()
        return

    if action == 'get_custom_ai_price':
        price = update.message.text
        if not price.isdigit():
            await update.message.reply_text("خطا: لطفاً فقط عدد وارد کنید.")
            return
        context.user_data['quote_price'] = int(price)
        context.user_data['next_admin_action'] = 'get_custom_ai_details'
        await update.message.reply_text("قیمت ثبت شد. ✅\n\nاکنون **توضیحات تکمیلی** (ویژگی‌ها، زمان تحویل و...) را برای ارسال به کاربر وارد کنید:")
        return

    if action == 'get_custom_ai_details':
        details = update.message.text
        order_id = context.user_data['target_order_id']
        price = context.user_data['quote_price']
        
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM orders WHERE order_id = ?", (order_id,))
        result = cursor.fetchone()
        if result:
            user_id = result[0]
            # Merge admin-provided price/details into existing crm_details_json while preserving user data
            cursor.execute("SELECT crm_details_json FROM orders WHERE order_id = ?", (order_id,))
            existing_row = cursor.fetchone()
            existing_crm = existing_row[0] if existing_row and existing_row[0] is not None else None
            merged = {}
            if existing_crm:
                try:
                    parsed = json.loads(existing_crm)
                    if isinstance(parsed, dict):
                        merged.update(parsed)
                    else:
                        # if stored as plain text, treat as user description
                        merged['description'] = str(parsed)
                except Exception:
                    merged['description'] = existing_crm

            # store admin fields under distinct keys to avoid clobbering user-provided fields
            merged['price'] = price
            merged['admin_details'] = details

            cursor.execute("UPDATE orders SET status = ?, total_price = ?, crm_details_json = ? WHERE order_id = ?", 
                           ('pending_user_approval', price, json.dumps(merged, ensure_ascii=False), order_id))
            conn.commit()
            try:
                log_order_event(order_id, f"ADMIN QUOTE SET price={price} details={details}")
                log_order_snapshot(order_id)
            except Exception:
                logger.exception("Failed to log admin quote for %s", order_id)
            try:
                log_order_event(order_id, f"ADMIN QUOTE SET price={price} details={details}")
                log_order_snapshot(order_id)
            except Exception:
                logger.exception("Failed to log admin quote for %s", order_id)

            user_message = (f"🔔 یک پیشنهاد برای سفارش کاستوم بات شما (`{order_id}`) ثبت شد:\n\n"
                            f"**💰 قیمت نهایی:** {persian_format_number(price)} تومان\n\n"
                            f"**📝 توضیحات ادمین:**\n{details}\n\n"
                            "آیا مایل به ادامه و پرداخت هستید؟")
            keyboard_user = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ بله، تایید و پرداخت", callback_data=f"user_confirm_custom_{order_id}")],
                [InlineKeyboardButton("❌ خیر، لغو سفارش", callback_data=f"user_reject_custom_{order_id}")]
            ])
            try:
                await context.bot.send_message(user_id, user_message, reply_markup=keyboard_user, parse_mode=ParseMode.MARKDOWN)
                await update.message.reply_text(f"✅ پیشنهاد با موفقیت برای کاربر ارسال شد.")
            except Exception as e:
                logger.error(f"Failed to send quote to user {user_id}: {e}")
                await update.message.reply_text("⚠️ خطا در ارسال پیام به کاربر.")
        else:
            await update.message.reply_text("خطا: سفارش یافت نشد.")
        
        conn.close()
        context.user_data.clear()
        return
    
    if action == 'get_user_rejection_reply':
        reply_text = update.message.text
        target_user_id = context.user_data['target_user_id_for_reply']
        try:
            await context.bot.send_message(target_user_id, f"پاسخ ادمین به پیام شما:\n\n{reply_text}")
            await update.message.reply_text("✅ پاسخ شما برای کاربر ارسال شد.")
        except Exception as e:
            logger.error(f"Failed to send admin reply to {target_user_id}: {e}")
            await update.message.reply_text("⚠️ خطا در ارسال پیام به کاربر.")
        context.user_data.clear()
        return

    # --- END: منطق جدید ---

    if action == 'get_quote_price':
        price = update.message.text
        if not price.isdigit():
            await update.message.reply_text("خطا: لطفاً فقط عدد وارد کنید.")
            return
        context.user_data['quote_price'] = int(price)
        context.user_data['next_admin_action'] = 'get_quote_details'
        await update.message.reply_text("قیمت ثبت شد. ✅\n\nاکنون **توضیحات تکمیلی** (ویژگی‌ها، زمان تحویل و...) را برای ارسال به کاربر وارد کنید:")
        return

    if action == 'get_quote_details':
        details = update.message.text
        order_id = context.user_data['target_order_id']
        price = context.user_data['quote_price']
        
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM orders WHERE order_id = ?", (order_id,))
        result = cursor.fetchone()
        if result:
            user_id = result[0]
            # Merge admin quote into existing crm_details_json preserving user-provided description
            cursor.execute("SELECT crm_details_json FROM orders WHERE order_id = ?", (order_id,))
            existing_row = cursor.fetchone()
            existing_crm = existing_row[0] if existing_row and existing_row[0] is not None else None
            merged = {}
            if existing_crm:
                try:
                    parsed = json.loads(existing_crm)
                    if isinstance(parsed, dict):
                        merged.update(parsed)
                    else:
                        merged['description'] = str(parsed)
                except Exception:
                    merged['description'] = existing_crm

            merged['price'] = price
            merged['admin_details'] = details

            cursor.execute("UPDATE orders SET status = ?, total_price = ?, crm_details_json = ? WHERE order_id = ?", 
                           ('pending_user_approval', price, json.dumps(merged, ensure_ascii=False), order_id))
            conn.commit()

            user_message = (f"🔔 یک پیشنهاد برای سفارش چت بات شما (`{order_id}`) ثبت شد:\n\n"
                            f"**💰 قیمت نهایی:** {persian_format_number(price)} تومان\n\n"
                            f"**📝 توضیحات ادمین:**\n{details}\n\n"
                            "آیا مایل به ادامه و پرداخت هستید؟")
            keyboard_user = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ بله، تایید و پرداخت", callback_data=f"user_confirm_{order_id}")],
                [InlineKeyboardButton("❌ خیر، لغو سفارش", callback_data=f"user_reject_{order_id}")]
            ])
            try:
                await context.bot.send_message(user_id, user_message, reply_markup=keyboard_user, parse_mode=ParseMode.MARKDOWN)
                await update.message.reply_text(f"✅ پیشنهاد با موفقیت برای کاربر ارسال شد.")
            except Exception as e:
                logger.error(f"Failed to send quote to user {user_id}: {e}")
                await update.message.reply_text("⚠️ خطا در ارسال پیام به کاربر.")

        else:
            await update.message.reply_text("خطا: سفارش یافت نشد.")
        
        conn.close()
        del context.user_data['next_admin_action']
        del context.user_data['target_order_id']
        del context.user_data['quote_price']
        return

    if action == 'get_bot_username':
        username = update.message.text
        context.user_data['bot_completion_username'] = username
        context.user_data['next_admin_action'] = 'get_bot_instructions'
        await update.message.reply_text("عالی. اکنون **توضیحات استفاده** از ربات را برای مشتری وارد کنید:", parse_mode=ParseMode.MARKDOWN)
        return

    if action == 'get_bot_instructions':
        instructions = update.message.text
        order_id = context.user_data['target_order_id']
        username = context.user_data['bot_completion_username']
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, full_name, phone, products_json, order_type FROM orders WHERE order_id = ?", (order_id,))
        result = cursor.fetchone()
        if result:
            user_id, full_name, phone, products_json_str, order_type = result
            products = json.loads(products_json_str)
            product_name = products[0]['name']

            cursor.execute("UPDATE orders SET status = 'completed' WHERE order_id = ?", (order_id,))
            conn.commit()
            try:
                user_final_message = (f"🎉 سفارش شما برای **{product_name}** آماده است! 🎉\n\n"
                                      f"**نام کاربری:** `{username}`\n\n"
                                      f"**توضیحات:**\n{instructions}")
                await context.bot.send_message(chat_id=user_id, text=user_final_message, parse_mode=ParseMode.MARKDOWN)
                await update.message.reply_text(f"✅ سفارش {order_id} با موفقیت تکمیل و اطلاعات برای مشتری ارسال شد.")
                
                product_id = get_product_id_by_name(product_name) or get_product_id_by_name("کاستوم بات هوش مصنوعی")
                
                if order_type in ['ai_credit', 'chatbot_request', 'custom_ai_request']:
                    expiry_date = (datetime.now() + timedelta(days=3650)).date().isoformat() # Permanent
                    log_to_excel(full_name, phone, product_name, order_id, "تکمیل شده", "دائمی")
                else: # For CRM bot
                    expiry_date = (datetime.now() + timedelta(days=30)).date().isoformat()
                    log_to_excel(full_name, phone, product_name, order_id, "تکمیل شده", 30)

                if product_id > 0:
                    cursor.execute(
                        "INSERT INTO active_services (user_id, product_id, product_name, expiry_date) VALUES (?, ?, ?, ?)",
                        (user_id, product_id, product_name, expiry_date)
                    )
                    conn.commit()
                    logger.info(f"Service/Credit '{product_name}' for user {user_id} added to active_services.")

            except Exception as e:
                logger.error(f"Failed to send completion message to user {user_id}: {e}")
                await update.message.reply_text(f"⚠️ **اخطار:** وضعیت سفارش {order_id} در دیتابیس به‌روز شد، اما ارسال پیام به کاربر ناموفق بود.")
        else:
            await update.message.reply_text(f"❌ خطای غیرمنتظره: سفارش {order_id} یافت نشد.")
        conn.close()
        del context.user_data['next_admin_action']
        del context.user_data['target_order_id']
        del context.user_data['bot_completion_username']

# --- توابع جدید برای مدیریت تایید/رد کاربر ---
async def user_confirm_custom_quote(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    order_id = "_".join(query.data.split("_")[3:])
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("SELECT total_price, full_name, phone FROM orders WHERE order_id = ?", (order_id,))
    result = cursor.fetchone()
    if not result:
        await query.edit_message_text("خطا: سفارش یافت نشد.")
        conn.close()
        return ConversationHandler.END
    
    price, full_name, phone = result
    cursor.execute("UPDATE orders SET status = ? WHERE order_id = ?", ('user_approved', order_id))
    conn.commit()
    conn.close()

    context.user_data['order_id_for_receipt'] = order_id
    
    await query.edit_message_text(
        f"پیشنهاد تایید شد. اطلاعات شما:\n"
        f" - نام: {full_name}\n"
        f" - تلفن: {phone}\n\n"
        f"لطفاً مبلغ **{persian_format_number(price)} تومان** را به شماره کارت زیر واریز و **عکس فیش** را ارسال کنید:\n\n`{BANK_CARD_NUMBER}`",
        parse_mode=ParseMode.MARKDOWN
    )
    return WAITING_FOR_CUSTOM_AI_RECEIPT

async def user_reject_custom_quote(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    order_id = "_".join(query.data.split("_")[3:])
    context.user_data['rejected_order_id'] = order_id

    keyboard = [
        [InlineKeyboardButton("✅ بله، مایلم", callback_data="reject_reason_yes")],
        [InlineKeyboardButton("❌ خیر", callback_data="reject_reason_no")]
    ]
    await query.edit_message_text(
        "شما پیشنهاد را رد کردید. آیا مایلید دلیل آن را برای ادمین ارسال کنید؟",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return USER_REJECT_REASON_PROMPT

async def user_reject_reason_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    choice = query.data.split('_')[-1]
    order_id = context.user_data['rejected_order_id']

    if choice == 'yes':
        await query.edit_message_text("لطفاً دلیل خود را در یک پیام بنویسید:")
        return USER_REJECT_GET_REASON
    else:
        conn = sqlite3.connect("shop_data.db", check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("UPDATE orders SET status = 'rejected_by_user' WHERE order_id = ?", (order_id,))
        conn.commit()
        conn.close()
        await query.edit_message_text("درخواست شما لغو گردید.")
        await context.bot.send_message(OWNER_ID, f"❌ کاربر پیشنهاد برای سفارش `{order_id}` را بدون ارائه دلیل رد کرد.")
        context.user_data.clear()
        return ConversationHandler.END

async def user_get_rejection_reason(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    reason = update.message.text
    order_id = context.user_data['rejected_order_id']
    user = update.effective_user

    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute("UPDATE orders SET status = 'rejected_by_user' WHERE order_id = ?", (order_id,))
    conn.commit()
    conn.close()

    await update.message.reply_text("دلیل شما برای ادمین ارسال شد. از بازخورد شما متشکریم.")
    
    admin_message = (f"❌ کاربر پیشنهاد برای سفارش `{order_id}` را رد کرد.\n\n"
                     f"**دلیل کاربر:**\n{reason}")
    
    keyboard = [[InlineKeyboardButton("💬 پاسخ به کاربر", callback_data=f"admin_reply_rejection_{user.id}")]]
    await context.bot.send_message(OWNER_ID, admin_message, reply_markup=InlineKeyboardMarkup(keyboard))
    
    context.user_data.clear()
    return ConversationHandler.END

async def admin_reply_to_rejection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = int(query.data.split('_')[-1])
    context.user_data['next_admin_action'] = 'get_user_rejection_reply'
    context.user_data['target_user_id_for_reply'] = user_id
    await query.message.reply_text(f"در حال پاسخ به کاربر با آیدی `{user_id}`. لطفاً پیام خود را ارسال کنید:", parse_mode=ParseMode.MARKDOWN)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "عملیات لغو شد. به منوی اصلی بازگشتید.",
        reply_markup=ReplyKeyboardRemove()
    )
    context.user_data.clear()
    await start(update, context)
    return ConversationHandler.END

async def check_subscriptions(context: ContextTypes.DEFAULT_TYPE):
    logger.info("Job: Checking for expired subscriptions...")
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    today_str = datetime.now().date().isoformat()
    
    cursor.execute(
        "SELECT id, user_id, product_name FROM active_services WHERE expiry_date < ? AND is_active = 1",
        (today_str,)
    )
    expired_services = cursor.fetchall()

    for service_id, user_id, product_name in expired_services:
        try:
            keyboard = [[InlineKeyboardButton(f"🔄 تمدید «{product_name}»", callback_data=f"renew_{service_id}")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await context.bot.send_message(
                chat_id=user_id,
                text=f"⚠️ اشتراک شما برای سرویس **{product_name}** به پایان رسیده است. برای استفاده مجدد، لطفاً آن را تمدید کنید.",
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=reply_markup
            )
            cursor.execute("UPDATE active_services SET is_active = 0 WHERE id = ?", (service_id,))
            logger.info(f"Notified user {user_id} about expired service {product_name} and deactivated it.")
        except Exception as e:
            logger.error(f"Failed to notify user {user_id} about expired service: {e}")
            
    conn.commit()
    conn.close()

async def my_services(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    user_id = update.effective_user.id
    
    if query: 
        await query.answer()
    
    conn = sqlite3.connect("shop_data.db", check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, product_name, expiry_date FROM active_services WHERE user_id = ? AND is_active = 1",
        (user_id,)
    )
    services = cursor.fetchall()
    conn.close()

    if not services:
        text = "شما در حال حاضر هیچ سرویس یا اعتبار فعالی ندارید."
        keyboard = [[InlineKeyboardButton("🛍️ مشاهده محصولات", callback_data="view_products")]]
    else:
        text = "👤 **سرویس‌های فعال و اعتبارهای شما:**\n\n"
        keyboard = []
        today = datetime.now().date()
        for service_id, name, expiry_str in services:
            if 'اعتبار چت جی پی تی' in name or 'چت بات' in name or 'کاستوم بات' in name:
                text += f"💎 **{name}**\n"
                text += f"   - نوع: سرویس دائمی/بسته اعتباری\n\n"
            else:
                expiry_date = datetime.strptime(expiry_str, '%Y-%m-%d').date()
                remaining_days = (expiry_date - today).days
                if remaining_days < 0: remaining_days = 0
                
                text += f"🔹 **{name}**\n"
                text += f"   - 🗓️ روزهای باقی‌مانده: **{remaining_days}** روز\n\n"
                keyboard.append([InlineKeyboardButton(f"🔄 تمدید سرویس «{name}»", callback_data=f"renew_{service_id}")])
    
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="main_menu")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if query:
        await query.edit_message_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)
    else:
        await update.message.reply_text(text, reply_markup=reply_markup, parse_mode=ParseMode.MARKDOWN)

    return MAIN_MENU
    
async def get_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id != OWNER_ID:
        return

    if os.path.exists(EXCEL_FILE_NAME):
        await update.message.reply_document(document=open(EXCEL_FILE_NAME, 'rb'))
    else:
        await update.message.reply_text("فایل اکسل هنوز ساخته نشده است. اولین سفارش که تایید شود، فایل ساخته خواهد شد.")

async def support(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"برای ارتباط با پشتیبانی، به آیدی زیر پیام دهید:\n{SUPPORT_USERNAME}")

def main() -> None:
    setup_database()
    setup_excel_file()
    application = Application.builder().token(BOT_TOKEN).build()
    
    job_queue = application.job_queue
    job_queue.run_daily(check_subscriptions, time=dt_time(hour=9, minute=0))

    # --- START OF FIX ---
    # مکالمه‌ی پرداخت برای چت‌بات (جداگانه برای جلوگیری از تداخل)
    chatbot_payment_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(handle_chatbot_quote_confirm, pattern=r"^user_confirm_(?!custom_)"),
            CallbackQueryHandler(handle_chatbot_quote_reject, pattern=r"^user_reject_(?!custom_)")
        ],
        states={
            WAITING_FOR_CHATBOT_RECEIPT: [MessageHandler(filters.PHOTO, handle_chatbot_payment_receipt)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )
    # --- END OF FIX ---

    # مکالمه جدید برای کاستوم بات هوش مصنوعی
    custom_ai_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(custom_ai_start, pattern="^custom_ai_start$")],
        states={
            CUSTOM_AI_START: [CallbackQueryHandler(custom_ai_ask_file, pattern="^custom_ai_type_")],
            CUSTOM_AI_ASK_FILE: [CallbackQueryHandler(custom_ai_get_file_choice, pattern="^custom_ai_has_file_")],
            CUSTOM_AI_GET_FILE: [MessageHandler(filters.Document.ALL, custom_ai_get_file)],
            CUSTOM_AI_GET_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, custom_ai_get_desc)],
            CUSTOM_AI_GET_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, custom_ai_get_name)],
            CUSTOM_AI_GET_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, custom_ai_finalize_request)],
        },
        fallbacks=[
            CallbackQueryHandler(show_product_details, pattern=f"^details_{get_product_id_by_name('بات هوش مصنوعی')}$"),
            CommandHandler("start", start)
        ],
        per_message=False
    )
    
    # مکالمه برای تایید/رد قیمت توسط کاربر
    user_quote_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(user_confirm_custom_quote, pattern="^user_confirm_custom_"),
            CallbackQueryHandler(user_reject_custom_quote, pattern="^user_reject_custom_")
        ],
        states={
            WAITING_FOR_CUSTOM_AI_RECEIPT: [MessageHandler(filters.PHOTO, receive_custom_ai_receipt)],
            USER_REJECT_REASON_PROMPT: [CallbackQueryHandler(user_reject_reason_prompt, pattern="^reject_reason_")],
            USER_REJECT_GET_REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_get_rejection_reason)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )

    # مکالمه اصلی که شامل تمام فرآیندهای دیگر است
    main_conv = ConversationHandler(
        entry_points=[
            CommandHandler("start", start), 
            CallbackQueryHandler(start_renewal, pattern="^renew_"),
        ],
        states={
            MAIN_MENU: [
                CallbackQueryHandler(back_to_products, pattern="^view_products$"),
                CallbackQueryHandler(my_services, pattern="^my_services$"),
                CallbackQueryHandler(start, pattern="^main_menu$"),
            ],
            SELECTING_PRODUCT: [
                CallbackQueryHandler(show_product_details, pattern="^details_"),
                CallbackQueryHandler(view_cart, pattern="^view_cart$"),
                CallbackQueryHandler(start, pattern="^main_menu$"),
                CallbackQueryHandler(checkout, pattern="^checkout$"),
                CallbackQueryHandler(remove_from_cart, pattern="^remove_")
            ],
            SHOWING_DETAILS: [
                CallbackQueryHandler(product_selection, pattern="^add_"),
                CallbackQueryHandler(view_cart, pattern="^view_cart$"),
                CallbackQueryHandler(back_to_products, pattern="^back_to_products$"),
                CallbackQueryHandler(prompt_ai_credit_options, pattern="^buy_credits_ai$"),
                CallbackQueryHandler(show_product_details, pattern="^details_"),
                CallbackQueryHandler(chatbot_platform_selected, pattern="^chatbot_")
            ],
            ASK_FOR_WHOM: [CallbackQueryHandler(ask_for_whom)],
            GET_TARGET_USER_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_target_user_id)],
            WAITING_FOR_RECEIPT: [MessageHandler(filters.PHOTO, receive_receipt)],
            AI_CREDIT_MENU: [
                CallbackQueryHandler(start_credit_purchase, pattern="^select_credit_"),
                CallbackQueryHandler(show_product_details, pattern="^details_")
            ],
            GET_CREDIT_BUYER_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_credit_buyer_name)],
            GET_CREDIT_BUYER_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_credit_buyer_phone)],
            WAITING_FOR_CREDIT_RECEIPT: [MessageHandler(filters.PHOTO, receive_credit_receipt)],
            ASK_CRM_LOGO: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_name_for_order)],
            ASK_CRM_SHOP_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, receive_phone_for_order)],
            ASK_PRODUCT_INFO_METHOD: [MessageHandler(filters.PHOTO | filters.TEXT, ask_crm_logo)],
            AWAITING_PRODUCT_FILE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_shop_name)],
            ASK_CRM_PRODUCT_COUNT: [CallbackQueryHandler(ask_product_info_method)],
            ASK_CRM_PRODUCT_NAMES: [MessageHandler(filters.Document.ALL, awaiting_product_file)],
            ASK_CRM_PRICES: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_product_count)],
            ASK_CRM_SHIPPING_QUERY: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_shipping_query)],
            ASK_CRM_CARD_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_product_names)],
            ASK_CRM_FULL_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_prices)],
            ASK_CRM_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_card_number)],
            WAITING_FOR_RENEWAL_RECEIPT: [MessageHandler(filters.PHOTO, receive_renewal_receipt)],
            RENEW_CRM_CHOICE: [CallbackQueryHandler(renew_crm_choice)],
            AWAITING_RENEW_CRM_FILE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_crm_shop_name)],
            CHATBOT_PLATFORM: [
                CallbackQueryHandler(chatbot_platform_selected, pattern="^chatbot_"),
                CallbackQueryHandler(back_to_products, pattern="^back_to_products$")
            ],
            CHATBOT_HAS_BOT: [CallbackQueryHandler(chatbot_has_bot, pattern="^has_bot_")],
            CHATBOT_GET_TOKEN: [MessageHandler(filters.TEXT & ~filters.COMMAND, chatbot_get_token)],
            CHATBOT_GET_SITE_INFO: [MessageHandler(filters.TEXT & ~filters.COMMAND, chatbot_get_site_info)],
            CHATBOT_GET_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, chatbot_get_desc)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False
    )

    application.add_handler(main_conv)
    application.add_handler(custom_ai_conv) 
    application.add_handler(user_quote_conv) 
    application.add_handler(chatbot_payment_conv) 
    
    # کنترل‌کننده‌هایی که باید همیشه فعال باشند
    application.add_handler(CommandHandler("admin", admin_pannel))
    application.add_handler(CommandHandler("getexcel", get_excel))
    application.add_handler(CommandHandler("cart", view_cart))
    application.add_handler(CommandHandler("my_services", my_services))
    application.add_handler(CommandHandler("support", support))
    application.add_handler(CallbackQueryHandler(admin_actions, pattern="^admin_"))
    application.add_handler(CallbackQueryHandler(admin_reply_to_rejection, pattern="^admin_reply_rejection_"))
    # --- START OF FIX ---
    # The problematic handler is now moved inside chatbot_payment_conv
    # --- END OF FIX ---
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_message_handler))

    print("ربات با قابلیت جدید (کاستوم بات هوش مصنوعی) در حال اجراست...")
    application.run_polling()


if __name__ == "__main__":
    main()
